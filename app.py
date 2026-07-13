import streamlit as st
import pandas as pd
import os
import io

try:
    from geo_processor import generate_district_assignment_from_shapes, generate_city_assignment_from_shapes, extract_precinct_metrics
    GEO_AVAILABLE = True
except ImportError:
    GEO_AVAILABLE = False

from main import run_pipeline, generate_template
import file_manager

print("[DEBUG] app.py execution cycle triggered.")
file_manager.sync_metadata_with_disk()

st.set_page_config(
    page_title="Priority Precinct Generator (Strict Mode)",
    page_icon="🗺️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- CSS Styling ---
css = """
<style>
.stApp { background-color: #f8fafc; }
.metric-box { background-color: white; padding: 20px; border-radius: 10px; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1); text-align: center; }
.metric-title { font-size: 14px; color: #64748b; text-transform: uppercase; letter-spacing: 1px; }
.metric-value { font-size: 32px; font-weight: bold; color: #0f172a; }
.status-box { background-color: white; padding: 15px; border-left: 5px solid #3b82f6; border-radius: 5px; margin-bottom: 20px;}
.hint-text { font-size: 13px; color: #6b7280; font-style: italic; margin-top: -10px; margin-bottom: 10px;}
</style>
"""
st.markdown(css, unsafe_allow_html=True)
os.makedirs("data", exist_ok=True)
os.makedirs("outputs", exist_ok=True)

def get_active_contest_file():
    for ext in ['.csv', '.tsv', '.xlsx', '.xls']:
        path = f"data/contest_data_input{ext}"
        if os.path.exists(path):
            return path
    return None

import zipfile
import json
from datetime import datetime
from pathlib import Path

def get_file_mime_type(path):
    ext = os.path.splitext(path)[1].lower()
    mime_types = {
        '.csv': 'text/csv',
        '.md': 'text/markdown',
        '.json': 'application/json',
        '.zip': 'application/zip',
        '.txt': 'text/plain'
    }
    return mime_types.get(ext, 'application/octet-stream')

def build_output_manifest(run_context=None):
    candidates = [
        # 1. Main Campaign Outputs
        {
            "label": "Production Priority Precincts CSV",
            "file_path": "outputs/final_rankings/production_priority_precincts.csv",
            "file_type": ".csv",
            "required_or_optional": "required",
            "category": "1. Main Campaign Outputs",
            "description": "Final prioritized precinct list for field operations."
        },
        {
            "label": "Base Preview Rankings CSV",
            "file_path": "outputs/final_rankings/base_preview_rankings.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "1. Main Campaign Outputs",
            "description": "Initial baseline rankings prior to contest data enrichment."
        },
        # 2. Top 50 and Explainability
        {
            "label": "Top 50 Explainability Table",
            "file_path": "outputs/final_rankings/top_50_explainability_table.csv",
            "file_type": ".csv",
            "required_or_optional": "required",
            "category": "2. Top 50 and Explainability",
            "description": "Detailed explanation of scores and metrics for the top 50 precincts."
        },
        {
            "label": "Rank Shift Report CSV",
            "file_path": "outputs/final_rankings/rank_shift_report.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "2. Top 50 and Explainability",
            "description": "Precinct-by-precinct comparison of base vs final rankings."
        },
        # 3. Validation Reports
        {
            "label": "Final Validation Summary",
            "file_path": "outputs/final_validation/final_validation_summary.md",
            "file_type": ".md",
            "required_or_optional": "required",
            "category": "3. Validation Reports",
            "description": "Main summary report of the production readiness validation."
        },
        {
            "label": "Contest Scope Validation",
            "file_path": "outputs/final_validation/contest_scope_validation.md",
            "file_type": ".md",
            "required_or_optional": "required",
            "category": "3. Validation Reports",
            "description": "Detailed scope validation report matching universe to contest boundaries."
        },
        {
            "label": "Final Config Reconciliation Verdict",
            "file_path": "outputs/contest_enrichment_reconciliation/final_config_reconciliation_verdict.md",
            "file_type": ".md",
            "required_or_optional": "required",
            "category": "3. Validation Reports",
            "description": "Verdict of the final contest classification configuration reconciliation."
        },
        {
            "label": "Architecture Alignment Report",
            "file_path": "outputs/architecture_alignment_report.md",
            "file_type": ".md",
            "required_or_optional": "optional",
            "category": "3. Validation Reports",
            "description": "Overall report mapping data structures and realigned precinct logic."
        },
        # 4. Crosswalk Files
        {
            "label": "Official Crosswalk Audit",
            "file_path": "outputs/precinct_crosswalk/canonical_sov_to_voter_precinct_crosswalk.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "4. Crosswalk Files",
            "description": "Canonical mapping of Voting Precincts (SOV) to Regular Precincts (Voter File)."
        },
        {
            "label": "Crosswalk Validation Summary",
            "file_path": "outputs/precinct_crosswalk/crosswalk_validation_summary.md",
            "file_type": ".md",
            "required_or_optional": "optional",
            "category": "4. Crosswalk Files",
            "description": "Validation summary for precinct crosswalk mapping."
        },
        {
            "label": "Crosswalk Match Audit",
            "file_path": "outputs/precinct_crosswalk/crosswalk_match_audit.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "4. Crosswalk Files",
            "description": "Detailed match audit logs for each precinct in the crosswalk."
        },
        {
            "label": "Crosswalk Coverage Simulation",
            "file_path": "outputs/precinct_crosswalk/crosswalk_coverage_simulation.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "4. Crosswalk Files",
            "description": "Simulation log verifying voter coverage under crosswalk mapping."
        },
        {
            "label": "Regular-to-Voting Xref (ewmr010)",
            "file_path": "outputs/precinct_crosswalk/parsed_regular_vbm_voting_xref.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "4. Crosswalk Files",
            "description": "Parsed regular precinct to VBM/voting precinct crosswalk log."
        },
        {
            "label": "Voting-to-Regular Xref (ewmr008)",
            "file_path": "outputs/precinct_crosswalk/parsed_voting_vbm_regular_xref.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "4. Crosswalk Files",
            "description": "Parsed voting precinct to VBM/regular precinct crosswalk log."
        },
        # 5. Diagnostic Files
        {
            "label": "Readiness Contradiction Report",
            "file_path": "outputs/final_validation/readiness_contradiction_report.md",
            "file_type": ".md",
            "required_or_optional": "optional",
            "category": "5. Diagnostic Files",
            "description": "Details on any contradictions in validation status."
        },
        {
            "label": "Proof Exports Summary",
            "file_path": "outputs/final_validation/proof_exports_summary.md",
            "file_type": ".md",
            "required_or_optional": "optional",
            "category": "5. Diagnostic Files",
            "description": "High-level summary of all generated proof exports."
        },
        {
            "label": "Contest Coverage Summary",
            "file_path": "outputs/final_validation/contest_coverage_summary.md",
            "file_type": ".md",
            "required_or_optional": "optional",
            "category": "5. Diagnostic Files",
            "description": "Detailed coverage analysis across all voter file precincts."
        },
        {
            "label": "Configuration Truth Report",
            "file_path": "outputs/final_validation/configuration_truth_report.md",
            "file_type": ".md",
            "required_or_optional": "optional",
            "category": "5. Diagnostic Files",
            "description": "Analysis of configuration values compared to database state."
        },
        {
            "label": "Context Consistency Report",
            "file_path": "outputs/final_validation/context_consistency_report.md",
            "file_type": ".md",
            "required_or_optional": "optional",
            "category": "5. Diagnostic Files",
            "description": "Verification report for context parameters consistency."
        },
        {
            "label": "Active Overrides Log",
            "file_path": "outputs/final_validation/active_overrides_log.json",
            "file_type": ".json",
            "required_or_optional": "optional",
            "category": "5. Diagnostic Files",
            "description": "JSON log listing all active validation overrides."
        },
        {
            "label": "Contest Scoring Breakdown",
            "file_path": "outputs/contest_data_manager/contest_scoring_breakdown.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "5. Diagnostic Files",
            "description": "Precinct-by-precinct scoring breakdown for active contests."
        },
        {
            "label": "Contest Rank Shift Report",
            "file_path": "outputs/contest_data_manager/contest_rank_shift_report.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "5. Diagnostic Files",
            "description": "Report tracking priority changes with contest enrichment."
        },
        {
            "label": "Contest Coverage Report",
            "file_path": "outputs/contest_data_manager/contest_coverage_report.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "5. Diagnostic Files",
            "description": "Report on voter-level contest coverage."
        },
        {
            "label": "Precinct Normalization Audit",
            "file_path": "outputs/contest_data_manager/precinct_normalization_audit.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "5. Diagnostic Files",
            "description": "Normalization trace log mapping raw precinct IDs."
        },
        {
            "label": "Architecture Alignment File Trace",
            "file_path": "outputs/architecture_alignment_file_trace.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "5. Diagnostic Files",
            "description": "Detailed file path trace mapping pipeline dependencies."
        },
        # 6. Contest Signal Model Outputs
        {
            "label": "Current Campaign Profile JSON",
            "file_path": "outputs/contest_signal_model/current_campaign_profile.json",
            "file_type": ".json",
            "required_or_optional": "optional",
            "category": "6. Contest Signal Model Outputs",
            "description": "JSON database containing the target candidate, cause, and campaign metadata."
        },
        {
            "label": "Contest Library JSON",
            "file_path": "outputs/contest_signal_model/contest_library.json",
            "file_type": ".json",
            "required_or_optional": "optional",
            "category": "6. Contest Signal Model Outputs",
            "description": "Historical contest data library tracking files, scopes, and weights."
        },
        {
            "label": "Contest Column Classification Matrix CSV",
            "file_path": "outputs/contest_signal_model/contest_column_classification_matrix.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "6. Contest Signal Model Outputs",
            "description": "Precinct-level historical column classification indicator mappings."
        },
        {
            "label": "Precinct Contest Signal Matrix CSV",
            "file_path": "outputs/contest_signal_model/precinct_contest_signal_matrix.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "6. Contest Signal Model Outputs",
            "description": "Calculated rate and vote counts per precinct for all contests."
        },
        {
            "label": "Aggregate Precinct Signal Scores CSV",
            "file_path": "outputs/contest_signal_model/aggregate_precinct_signal_scores.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "6. Contest Signal Model Outputs",
            "description": "Averaged support, opposition, margin, and turnout signal scores."
        },
        {
            "label": "Preview Multi-Contest Priority Scores CSV",
            "file_path": "outputs/contest_signal_model/preview_multi_contest_priority_scores.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "6. Contest Signal Model Outputs",
            "description": "Preview precinct priority scores, rankings, rank changes, and strategic planning buckets."
        },
        {
            "label": "Contest Signal Correlation Matrix CSV",
            "file_path": "outputs/contest_signal_model/contest_signal_correlation_matrix.csv",
            "file_type": ".csv",
            "required_or_optional": "optional",
            "category": "6. Contest Signal Model Outputs",
            "description": "Pearson correlation statistics comparing contest indicators across precincts."
        },
        {
            "label": "Contest Signal Validation Report MD",
            "file_path": "outputs/contest_signal_model/contest_signal_validation_report.md",
            "file_type": ".md",
            "required_or_optional": "optional",
            "category": "6. Contest Signal Model Outputs",
            "description": "Markdown validation diagnostic summary."
        }
    ]
    
    manifest = []
    for item in candidates:
        path = item["file_path"]
        exists = os.path.exists(path)
        item["exists"] = exists
        item["size_bytes"] = os.path.getsize(path) if exists else None
        manifest.append(item)
    return manifest

def persist_output_manifest(manifest):
    st.session_state["latest_output_manifest"] = manifest

def create_outputs_zip(manifest, run_context=None):
    os.makedirs("outputs/final_downloads", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_path = Path(f"outputs/final_downloads/ppg_validation_outputs_{timestamp}.zip")
    
    files_included = []
    files_missing = []
    
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for item in manifest:
            src_path = Path(item["file_path"])
            if item["exists"] and src_path.exists():
                rel_zip_path = src_path.relative_to("outputs")
                zipf.write(src_path, arcname=rel_zip_path)
                files_included.append(str(rel_zip_path))
            else:
                files_missing.append(item["file_path"])
                
        # Generate download_manifest.json
        manifest_data = {
            "generated_timestamp": datetime.now().isoformat(),
            "run_mode": run_context.get("run_mode", "USER_DASHBOARD_MODE") if run_context else "USER_DASHBOARD_MODE",
            "active_voter_file": run_context.get("active_voter_file", "data/voter_file.csv") if run_context else "data/voter_file.csv",
            "active_contest_file": run_context.get("active_contest_file", "") if run_context else "",
            "active_cross_reference_files": run_context.get("active_cross_reference_files", []) if run_context else [],
            "readiness_verdict": run_context.get("readiness_verdict", "unknown") if run_context else "unknown",
            "files_included": files_included,
            "files_missing": files_missing
        }
        
        zipf.writestr("download_manifest.json", json.dumps(manifest_data, indent=2))
        
    return zip_path

def render_download_panel(manifest, run_context=None):
    st.markdown("---")
    st.markdown("## 📥 Download Center")
    
    existing_items = [item for item in manifest if item["exists"]]
    if not existing_items:
        st.info("No files available for download yet. Run the pipeline to generate outputs.")
        return

    categories = [
        "1. Main Campaign Outputs",
        "2. Top 50 and Explainability",
        "3. Validation Reports",
        "4. Crosswalk Files",
        "5. Diagnostic Files",
        "6. Contest Signal Model Outputs"
    ]
    
    for category in categories:
        cat_items = [item for item in manifest if item["category"] == category and item["exists"]]
        if cat_items:
            st.markdown(f"### {category}")
            for item in cat_items:
                path = item["file_path"]
                label = item["label"]
                desc = item["description"]
                size_str = ""
                if item["size_bytes"] is not None:
                    size_kb = item["size_bytes"] / 1024.0
                    size_str = f" ({size_kb:.1f} KB)"
                
                col_info, col_btn = st.columns([3, 1])
                col_info.markdown(f"**{label}**{size_str}  \n*{desc}*")
                
                mime = get_file_mime_type(path)
                try:
                    with open(path, "rb") as f:
                        file_bytes = f.read()
                    col_btn.download_button(
                        label=f"📥 Download {os.path.basename(path)}",
                        data=file_bytes,
                        file_name=os.path.basename(path),
                        mime=mime,
                        key=f"dl_{path.replace('/', '_').replace('.', '_')}"
                    )
                except Exception as e:
                    col_btn.error(f"Error loading file: {e}")
                    
    st.markdown("### 6. Download Everything")
    col_info, col_btn = st.columns([3, 1])
    col_info.markdown("**Download All Outputs (.zip)**  \n*Packages all generated outputs and validation reports into a single ZIP archive.*")
    
    try:
        zip_path = create_outputs_zip(manifest, run_context)
        with open(zip_path, "rb") as f:
            zip_bytes = f.read()
        col_btn.download_button(
            label="📥 Download All Outputs (.zip)",
            data=zip_bytes,
            file_name=os.path.basename(zip_path),
            mime="application/zip",
            key="dl_all_outputs_zip"
        )
    except Exception as e:
        col_btn.error(f"Error creating ZIP: {e}")


def save_uploaded(file_obj, filename):
    if file_obj is not None:
        with open(os.path.join("data", filename), "wb") as f:
            f.write(file_obj.getbuffer())

# --- REAL-TIME CHECKER & DEPENDENCY MAP ---
def check_status():
    dist_exists = os.path.exists("data/district_assignment.csv")
    is_mock = False
    if dist_exists:
        try:
            df_check = pd.read_csv("data/district_assignment.csv", nrows=5)
            first_col = df_check.columns[0]
            is_mock = df_check[first_col].astype(str).str.contains("SRPREC_").any() or "mock" in dist_path.lower()
        except:
            is_mock = False

    status = {
        "voter": os.path.exists("data/voter_file.csv"),
        "mprec": os.path.exists("data/mprec_srprec.csv"),
        "city": os.path.exists("data/srprec_city.csv"),
        "dist": dist_exists and not is_mock,
        "is_mock_dist_present": dist_exists and is_mock,
        "metrics": os.path.exists("data/srprec_metrics.csv"),
        "has_prior_turnout": False,
        
        "shp_srprec": os.path.exists("data/srprec_shapes.zip"),
        "shp_city": os.path.exists("data/city_shapes.zip"),
        "shp_assem": os.path.exists("data/assembly_shapes.zip"),
        "shp_sup": os.path.exists("data/supervisorial_shapes.zip")
    }
    status["dist_shapefiles_ready"] = status["shp_srprec"] and status["shp_assem"] and status["shp_sup"]
    status["city_shapefiles_ready"] = status["shp_srprec"] and status["shp_city"]
    
    # Voter file alone enables Base Preview scoring
    status["ready_to_preview"] = status["voter"]
    
    # Check for Prior Turnout in Voter File safely
    if status["voter"]:
        try:
            head = pd.read_csv("data/voter_file.csv", nrows=1)
            cols = [c.lower().strip() for c in head.columns]
            if 'general22' in cols or any('2022' in c for c in cols):
                status["has_prior_turnout"] = True
        except:
            pass

    return status

status = check_status()

def get_file_mtime(path):
    if os.path.exists(path):
        return os.path.getmtime(path)
    return 0.0

@st.cache_data
def cached_inspect_and_load_file(file_path, sheet_name=None, mtime=0.0):
    import contest_manager
    return contest_manager.inspect_and_load_file(file_path, sheet_name=sheet_name)

@st.cache_data
def cached_generate_file_inventory(file_path, sheet_name=None, mtime=0.0):
    import contest_manager
    return contest_manager.generate_file_inventory(file_path, sheet_name=sheet_name)

@st.cache_data
def get_cached_match_rate(voter_file_path, v_mtime, contest_file_path, c_mtime, contest_prec_col):
    import contest_manager
    from main import to_clean_district_str
    voter_df = pd.read_csv(voter_file_path, low_memory=False)
    voter_precincts = voter_df['PrecinctName'].dropna().apply(to_clean_district_str).astype(str).str.strip().str.upper().unique().tolist()
    res_load = contest_manager.inspect_and_load_file(contest_file_path)
    if res_load["status"] == "success":
        match_res = contest_manager.generate_precinct_match_report(res_load["df"], contest_prec_col, voter_precincts)
        if match_res["status"] == "success":
            return match_res["match_rate"]
    return 0.0

@st.cache_data
def get_cached_precinct_match_report(voter_file_path, v_mtime, contest_file_path, selected_sheet, c_mtime, contest_prec_col):
    import contest_manager
    from main import to_clean_district_str
    voter_df = pd.read_csv(voter_file_path, low_memory=False)
    voter_precincts = voter_df['PrecinctName'].dropna().apply(to_clean_district_str).astype(str).str.strip().str.upper().unique().tolist()
    res_load = contest_manager.inspect_and_load_file(contest_file_path, sheet_name=selected_sheet)
    if res_load["status"] == "success":
        return contest_manager.generate_precinct_match_report(res_load["df"], contest_prec_col, voter_precincts)
    return {"status": "error", "message": "Failed to load contest file."}

# Cached geo option extractor for dropdowns and metadata
@st.cache_data
def get_cached_geo_options(voter_file_path, mprec_path, city_path, dist_path, derive_sonoma_sd, voter_col_mappings=None, mprec_col_mappings=None, city_col_mappings=None, dist_col_mappings=None):
    from main import find_voter_geo_columns, derive_sonoma_supervisorial, is_mock_district_file, to_clean_district_str
    
    ad_opts = ["ALL"]
    sd_opts = ["ALL"]
    city_opts = ["ALL"]
    metadata = {
        "supervisorial": {"source": "unmapped", "confidence": "unknown"},
        "assembly": {"source": "unmapped", "confidence": "unknown"},
        "senate": {"source": "unmapped", "confidence": "unknown"},
        "city": {"source": "unmapped", "confidence": "unknown"}
    }
    
    if not os.path.exists(voter_file_path):
        return ad_opts, sd_opts, city_opts, metadata
        
    try:
        voter_df = pd.read_csv(voter_file_path, low_memory=False)
    except Exception as e:
        return ad_opts, sd_opts, city_opts, metadata
        
    geo_cols = find_voter_geo_columns(voter_df, overrides=voter_col_mappings)
    
    # 1. City Options
    if geo_cols['city'] and geo_cols['city'] in voter_df.columns:
        cities = voter_df[geo_cols['city']].dropna().unique().tolist()
        city_opts += sorted([str(x) for x in cities])
        metadata['city'] = {"source": "voter_file_direct", "confidence": "high"}
    elif os.path.exists(city_path):
        try:
            city_df = pd.read_csv(city_path)
            c_city_col = city_col_mappings.get('city', 'city') if city_col_mappings else 'city'
            actual_c_city = next((c for c in city_df.columns if c.lower().strip() == c_city_col.lower().strip()), city_df.columns[1] if len(city_df.columns) > 1 else city_df.columns[0])
            cities = city_df[actual_c_city].dropna().unique().tolist()
            city_opts += sorted([str(x) for x in cities])
            metadata['city'] = {"source": "external_mapping", "confidence": "medium"}
        except:
            pass
            
    # 2. Assembly Options
    if geo_cols['assembly'] and geo_cols['assembly'] in voter_df.columns:
        ads = voter_df[geo_cols['assembly']].dropna().unique().tolist()
        ad_opts += sorted(list(set([to_clean_district_str(x) for x in ads if to_clean_district_str(x) != 'Unmapped'])))
        metadata['assembly'] = {"source": "voter_file_direct", "confidence": "high"}
    elif os.path.exists(dist_path):
        try:
            dist_df = pd.read_csv(dist_path)
            if not is_mock_district_file(dist_df, dist_path):
                d_assem_col = dist_col_mappings.get('assembly', 'assembly_district') if dist_col_mappings else 'assembly_district'
                actual_d_assem = next((c for c in dist_df.columns if c.lower().strip() == d_assem_col.lower().strip()), dist_df.columns[1] if len(dist_df.columns) > 1 else dist_df.columns[0])
                ads = dist_df[actual_d_assem].dropna().unique().tolist()
                ad_opts += sorted(list(set([to_clean_district_str(x) for x in ads if to_clean_district_str(x) != 'Unmapped'])))
                metadata['assembly'] = {"source": "external_mapping", "confidence": "medium"}
        except:
            pass
            
    # 3. Supervisorial Options
    if geo_cols['supervisorial'] and geo_cols['supervisorial'] in voter_df.columns:
        sds = voter_df[geo_cols['supervisorial']].dropna().unique().tolist()
        sd_opts += sorted(list(set([to_clean_district_str(x) for x in sds if to_clean_district_str(x) != 'Unmapped'])))
        metadata['supervisorial'] = {"source": "voter_file_direct", "confidence": "high"}
    elif derive_sonoma_sd:
        derived_sds = voter_df['PrecinctName'].dropna().apply(derive_sonoma_supervisorial).dropna().unique().tolist()
        sd_opts += sorted(list(set([to_clean_district_str(x) for x in derived_sds if to_clean_district_str(x) != 'Unmapped'])))
        metadata['supervisorial'] = {"source": "precinct_prefix_rule", "confidence": "high_sonoma_verified"}
    elif os.path.exists(dist_path):
        try:
            dist_df = pd.read_csv(dist_path)
            if not is_mock_district_file(dist_df, dist_path):
                d_sup_col = dist_col_mappings.get('supervisorial', 'supervisorial_district') if dist_col_mappings else 'supervisorial_district'
                actual_d_sup = next((c for c in dist_df.columns if c.lower().strip() == d_sup_col.lower().strip()), dist_df.columns[2] if len(dist_df.columns) > 2 else dist_df.columns[0])
                sds = dist_df[actual_d_sup].dropna().unique().tolist()
                sd_opts += sorted(list(set([to_clean_district_str(x) for x in sds if to_clean_district_str(x) != 'Unmapped'])))
                metadata['supervisorial'] = {"source": "external_mapping", "confidence": "medium"}
        except:
            pass
            
    return ad_opts, sd_opts, city_opts, metadata

st.title("🗺️ Priority Precinct Generator (Realignment Mode)")
st.caption("Voter-File First. Required Contests for Production. Optional Geography.")

# --- SIDEBAR: STRICT CONFIGURATION ---
with st.sidebar:
    st.header("⚙️ Score Configuration")
    
    can_underperf = status["has_prior_turnout"]
    can_density = status["metrics"] or status["shp_srprec"]

    st.subheader("Priority Formula Weights")
    
    weight_turnout = st.slider(
        "Turnout Opportunity", 0.0, 1.0, 0.45 if can_underperf else 0.0, 
        disabled=not can_underperf,
        help="Requires voters file with prior-cycle history (e.g. 2022 turnout)."
    )
    if not can_underperf: st.caption("❌ Unavailable: Missing Prior-Cycle Turnout")

    weight_comp = st.slider("Partisan Competitiveness", 0.0, 1.0, 0.35)
    
    # Geography indicator / proxy label
    if can_density:
        st.success("🟢 True Area Density Available")
        label = "True Area Density (Area)"
    else:
        st.info("ℹ️ Using Operational Scale Proxy")
        label = "Operational Scale Proxy"
        
    weight_density = st.slider(
        label, 0.0, 1.0, 0.20
    )

    tot = weight_turnout + weight_comp + weight_density
    if tot == 0:
        tot = 1
        weight_comp = 1.0
    
    weights = {
        "turnout_gap": weight_turnout/tot, 
        "competitive_index": weight_comp/tot, 
        "density": weight_density/tot
    }

    st.subheader("Turnout Opportunity Settings")
    election_context = st.selectbox(
        "Election Context",
        ["General", "Midterm", "Primary", "Special", "Other"],
        index=2
    )
    
    use_override = st.checkbox("Override Target Turnout", value=False)
    if use_override:
        target_turnout_override = st.slider("Target Turnout", 0.0, 1.0, 0.45)
    else:
        target_turnout_override = None
        
    enforce_size_guardrail = st.checkbox("Enforce Tiny Precinct Size Guardrail", value=True, help="Penalizes precincts with < 150 voters to avoid mathematical artifacts.")

    st.subheader("Sonoma County Configuration")
    is_sonoma = False
    if status["voter"]:
        from main import is_sonoma_context
        try:
            city_check = pd.read_csv("data/srprec_city.csv") if os.path.exists("data/srprec_city.csv") else None
        except:
            city_check = None
        is_sonoma = is_sonoma_context("data/voter_file.csv", city_check)
        
    derive_sonoma_sd = st.checkbox(
        "Derive Sonoma Supervisorial District from PrecinctName prefix",
        value=is_sonoma,
        help="Analyzes 6- or 7-digit precinct numbers to extract supervisor districts 1-5."
    )

    st.subheader("Contest Data Enrichment")
    active_contest_file = get_active_contest_file()
    contest_prec_col = None
    match_rate_val = 0.0
    config_count = 0
    if active_contest_file:
        col_file = "outputs/contest_data_manager/contest_precinct_col.txt"
        if os.path.exists(col_file):
            with open(col_file, "r") as f:
                contest_prec_col = f.read().strip()
        import contest_manager
        config_count = len(contest_manager.load_classification_config())
        if status["voter"] and contest_prec_col:
            try:
                v_mtime = get_file_mtime("data/voter_file.csv")
                c_mtime = get_file_mtime(active_contest_file)
                match_rate_val = get_cached_match_rate("data/voter_file.csv", v_mtime, active_contest_file, c_mtime, contest_prec_col)
            except:
                pass

    production_ranking_ready = status["voter"] and active_contest_file and contest_prec_col and (config_count > 0)

    if production_ranking_ready:
        st.success(f"✅ Production Contest File Active ({config_count} Contests)")
        st.markdown(f"- **Precinct Match Rate:** `{match_rate_val:.1f}%`")
        contest_influence_weight = st.slider(
            "Contest Data Influence Weight", 0.10, 0.60, 0.30,
            help="Determine the percentage contribution of the contest data to the final priority score."
        )
    else:
        st.info("ℹ️ Production scoring locked. Upload & classify contests to enable.")
        contest_influence_weight = 0.0

# --- STATUS PANEL ---
st.markdown('<div class="status-box">', unsafe_allow_html=True)
st.subheader("📊 Data Dependency Status")
stat_cols = st.columns(5)
stat_cols[0].markdown(f"**Voter File (Required):** {'✅ Loaded' if status['voter'] else '❌ Missing'}")
stat_cols[1].markdown(f"**Contest Data:** {'✅ Production Ranking Ready' if production_ranking_ready else '❌ Missing/Unclassified'}")
stat_cols[2].markdown(f"**Optional Crosswalk:** {'✅ Loaded' if status['mprec'] else '❌ Missing'}")
stat_cols[3].markdown(f"**Optional City Map:** {'✅ Loaded' if status['city'] else '❌ Missing'}")
stat_cols[4].markdown(f"**Optional Area Metrics:** {'✅ Loaded' if status['metrics'] else '❌ Missing'}")

if status['is_mock_dist_present']:
    st.warning("⚠️ Mock district_assignment.csv detected. Ignoring for production use.")
st.markdown('</div>', unsafe_allow_html=True)

# --- TABS ---
tab_file, tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📂 Central File Manager",
    "📁 1. Core Data Upload", 
    "🏙️ 2. City Mapping Manager (Optional)", 
    "🗺️ 3. District Mapping Manager (Optional)", 
    "📊 4. Contest Data Manager (Required for Production)",
    "🚀 5. Execution & Results",
    "📈 6. Contest Signal Manager"
])

# --- CENTRAL FILE MANAGER TAB ---
with tab_file:
    st.markdown("### 📂 Central File Manager")
    st.markdown("Tagging files binds them to specific system roles in the precinct prioritization model.")
    
    # 1. File Upload Section
    st.subheader("📤 Add Files")
    uploaded_files = st.file_uploader("Upload CSV, TSV, PDF, or ZIP shapefiles", accept_multiple_files=True, key="fm_uploader", label_visibility="collapsed")
    if uploaded_files:
        has_new = any(not os.path.exists(os.path.join("data", uf.name)) for uf in uploaded_files)
        if has_new:
            for uf in uploaded_files:
                dest = os.path.join("data", uf.name)
                if not os.path.exists(dest):
                    with open(dest, "wb") as f:
                        f.write(uf.getbuffer())
            file_manager.sync_metadata_with_disk()
            st.success("Files uploaded successfully!")
            st.rerun()
        
    st.markdown("---")
    
    metadata = file_manager.load_file_metadata()
    
    # 2. Active System Roles Table
    st.subheader("🎯 Active System Roles")
    
    active_mappings = {role: "❌ Unassigned" for role in file_manager.SYSTEM_ROLES.keys()}
    for fname, info in metadata.items():
        role = info.get("tag")
        if role and role in active_mappings:
            active_mappings[role] = fname
            
    role_rows = []
    for role, assigned in active_mappings.items():
        role_rows.append({
            "System Role": role,
            "Mapped File": assigned,
            "Status": "✅ Active" if assigned != "❌ Unassigned" else "⚠️ Missing"
        })
    st.dataframe(pd.DataFrame(role_rows), use_container_width=True, hide_index=True)
    
    st.markdown("---")
    
    # 3. User Files Section
    st.subheader("📄 Managed Files (Data Folder)")
    unarchived_files = {k: v for k, v in metadata.items() if not v.get("archived")}
    
    if not unarchived_files:
        st.info("No unarchived user files found. Upload some files above to get started!")
    else:
        c_h1, c_h2, c_h3, c_h4, c_h5 = st.columns([3, 2, 2, 2, 2])
        c_h1.markdown("**File Name & Details**")
        c_h2.markdown("**Tag/System Role**")
        c_h3.markdown("**Active Tag**")
        c_h4.markdown("**Archive Action**")
        c_h5.markdown("**Delete Action**")
        
        for fname, info in unarchived_files.items():
            path = os.path.join("data", fname)
            size_kb = os.path.getsize(path) / 1024.0 if os.path.exists(path) else 0.0
            
            c1, c2, c3, c4, c5 = st.columns([3, 2, 2, 2, 2])
            with c1:
                st.markdown(f"📁 **{fname}**")
                st.caption(f"Size: {size_kb:.1f} KB | Updated: {info.get('uploaded_at')}")
            with c2:
                available_roles = ["None (Untagged)"] + list(file_manager.SYSTEM_ROLES.keys())
                current_tag = info.get("tag")
                default_idx = available_roles.index(current_tag) if current_tag in available_roles else 0
                
                selected_role = st.selectbox(
                    f"Tag for {fname}",
                    available_roles,
                    index=default_idx,
                    key=f"tag_select_{fname}",
                    label_visibility="collapsed"
                )
                
                if selected_role == "None (Untagged)":
                    selected_role = None
                if selected_role != current_tag:
                    success, msg = file_manager.assign_tag_role(fname, selected_role)
                    if success:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
            with c3:
                if info.get("tag"):
                    st.success(f"🏷️ {info.get('tag')}")
                else:
                    st.info("No tag")
            with c4:
                if st.button("📦 Archive", key=f"arch_btn_{fname}", use_container_width=True):
                    success, msg = file_manager.archive_file(fname)
                    if success:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
            with c5:
                if st.button("🗑️ Delete", key=f"del_btn_{fname}", use_container_width=True):
                    success, msg = file_manager.delete_file(fname)
                    if success:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
                        
    st.markdown("---")
    
    # 4. Archived Files Section
    st.subheader("📦 Archived Files")
    archived_files = {k: v for k, v in metadata.items() if v.get("archived")}
    
    if not archived_files:
        st.info("No files currently archived.")
    else:
        c_ah1, c_ah2, c_ah3 = st.columns([6, 3, 3])
        c_ah1.markdown("**File Name & Details**")
        c_ah2.markdown("**Restore Action**")
        c_ah3.markdown("**Delete Action**")
        
        for fname, info in archived_files.items():
            path = os.path.join("data", "archive", fname)
            size_kb = os.path.getsize(path) / 1024.0 if os.path.exists(path) else 0.0
            
            c1, c2, c3 = st.columns([6, 3, 3])
            with c1:
                st.markdown(f"📦 *{fname}*")
                st.caption(f"Size: {size_kb:.1f} KB | Archived: {info.get('uploaded_at')}")
            with c2:
                if st.button("📤 Restore / Unarchive", key=f"unarch_btn_{fname}", use_container_width=True):
                    success, msg = file_manager.unarchive_file(fname)
                    if success:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
            with c3:
                if st.button("🗑️ Permanent Delete", key=f"del_arch_btn_{fname}", use_container_width=True):
                    success, msg = file_manager.delete_file(fname)
                    if success:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)

# --- TAB 1: CORE DATA ---
with tab1:
    st.markdown("### Upload Core Voter Files")
    st.info("💡 **Voter File First:** Aggregates database rows directly by `PrecinctName` as the core unit.")
    
    c1, c2 = st.columns(2)
    with c1:
        voter_up = st.file_uploader("1. Voter File (`voter_file.csv`)", type=['csv'])
        if voter_up:
            save_uploaded(voter_up, "voter_file.csv")
            st.success("✅ `voter_file.csv` has been successfully uploaded and saved!")
    with c2:
        mprec_up = st.file_uploader("2. MPREC Crosswalk (`mprec_srprec.csv`) [Optional]", type=['csv'])
        if mprec_up:
            save_uploaded(mprec_up, "mprec_srprec.csv")
            st.success("✅ `mprec_srprec.csv` has been successfully uploaded and saved!")
            
    st.markdown("---")
    st.markdown("#### 📁 Current Core Data Status on Disk")
    sc1, sc2 = st.columns(2)
    if os.path.exists("data/voter_file.csv"):
        v_size = os.path.getsize("data/voter_file.csv") / 1024.0
        sc1.success(f"🟢 **voter_file.csv** is active on disk ({v_size:.1f} KB)")
        if sc1.button("🗑️ Delete Voter File", key="del_voter_file"):
            os.remove("data/voter_file.csv")
            st.rerun()
    else:
        sc1.error("🔴 **voter_file.csv** is missing")
        
    if os.path.exists("data/mprec_srprec.csv"):
        m_size = os.path.getsize("data/mprec_srprec.csv") / 1024.0
        sc2.success(f"🟢 **mprec_srprec.csv** is active on disk ({m_size:.1f} KB)")
        if sc2.button("🗑️ Delete Crosswalk File", key="del_crosswalk_file"):
            os.remove("data/mprec_srprec.csv")
            st.rerun()
            
        try:
            mprec_cols = pd.read_csv("data/mprec_srprec.csv", nrows=1).columns.tolist()
        except:
            mprec_cols = []
            
        if mprec_cols:
            with st.expander("🗺️ Crosswalk File Column Mapping (Confirm / Override)", expanded=False):
                if "mprec_col_mappings" not in st.session_state:
                    m_cols_lower = {c.lower().strip(): c for c in mprec_cols}
                    m_def = m_cols_lower.get("mprec", mprec_cols[0] if mprec_cols else None)
                    s_def = m_cols_lower.get("srprec", mprec_cols[1] if len(mprec_cols) > 1 else mprec_cols[0])
                    st.session_state["mprec_col_mappings"] = {"mprec": m_def, "srprec": s_def}
                    
                m_maps = st.session_state["mprec_col_mappings"]
                m_maps["mprec"] = st.selectbox(
                    "MPREC Column (Voter File Precinct Key)",
                    mprec_cols,
                    index=mprec_cols.index(m_maps["mprec"]) if m_maps.get("mprec") in mprec_cols else 0
                )
                m_maps["srprec"] = st.selectbox(
                    "SRPREC Column (GIS Precinct Key)",
                    mprec_cols,
                    index=mprec_cols.index(m_maps["srprec"]) if m_maps.get("srprec") in mprec_cols else 0
                )
                st.session_state["resolved_mprec_mappings"] = m_maps
    else:
        sc2.warning("🟡 **mprec_srprec.csv** is missing (Bypassed; using direct PrecinctName)")
        
    if status["voter"]:
        try:
            voter_cols = pd.read_csv("data/voter_file.csv", nrows=1).columns.tolist()
        except:
            voter_cols = []
            
        if voter_cols:
            st.markdown("---")
            with st.expander("🗺️ Voter File Column Mapping (Confirm / Override)", expanded=False):
                st.markdown("Use these selectors to manually map columns in your voter file if they are not automatically detected.")
                
                if "voter_col_mappings" not in st.session_state:
                    try:
                        voter_df_sample = pd.read_csv("data/voter_file.csv", nrows=5)
                        from main import find_voter_geo_columns
                        defaults = find_voter_geo_columns(voter_df_sample)
                        
                        v_cols_lower = {c.lower().strip(): c for c in voter_cols}
                        p_def = v_cols_lower.get('precinctname', voter_cols[0] if voter_cols else None)
                        party_def = v_cols_lower.get('party', voter_cols[1] if len(voter_cols) > 1 else None)
                        t24_cols = [c for c in voter_cols if '24' in c or '2024' in c]
                        t24_def = t24_cols[0] if t24_cols else next((c for c in voter_cols if 'general' in c.lower()), None)
                        t22_cols = [c for c in voter_cols if '22' in c or '2022' in c.lower()]
                        t22_def = t22_cols[0] if t22_cols else None
                        
                        st.session_state["voter_col_mappings"] = {
                            "precinctname": p_def,
                            "party": party_def,
                            "turnout24": t24_def,
                            "turnout22": t22_def,
                            "supervisorial": defaults.get("supervisorial"),
                            "assembly": defaults.get("assembly"),
                            "senate": defaults.get("senate"),
                            "congressional": defaults.get("congressional"),
                            "city": defaults.get("city"),
                            "city_council": defaults.get("city_council"),
                            "school": defaults.get("school"),
                            "water": defaults.get("water"),
                            "special": defaults.get("special")
                        }
                    except:
                        st.session_state["voter_col_mappings"] = {}
                
                mappings = st.session_state["voter_col_mappings"]
                col_c1, col_c2 = st.columns(2)
                
                with col_c1:
                    st.markdown("##### Required Core Columns")
                    mappings["precinctname"] = st.selectbox(
                        "Precinct Column (PrecinctName)",
                        voter_cols,
                        index=voter_cols.index(mappings["precinctname"]) if mappings.get("precinctname") in voter_cols else 0
                    )
                    mappings["party"] = st.selectbox(
                        "Party Column",
                        voter_cols,
                        index=voter_cols.index(mappings["party"]) if mappings.get("party") in voter_cols else 0
                    )
                    mappings["turnout24"] = st.selectbox(
                        "Turnout 2024 / Current Column",
                        voter_cols,
                        index=voter_cols.index(mappings["turnout24"]) if mappings.get("turnout24") in voter_cols else 0
                    )
                    mappings["turnout22"] = st.selectbox(
                        "Turnout 2022 / Prior Column (Optional)",
                        ["None (Turnout Dropoff Disabled)"] + voter_cols,
                        index=(voter_cols.index(mappings["turnout22"]) + 1) if mappings.get("turnout22") in voter_cols else 0
                    )
                    
                with col_c2:
                    st.markdown("##### Optional Geographic Districts")
                    for key in ["supervisorial", "assembly", "senate", "congressional", "city", "city_council", "school", "water", "special"]:
                        label = key.capitalize() if key != "city_council" else "City Council District"
                        if key == "supervisorial": label = "Supervisorial District"
                        elif key == "assembly": label = "Assembly District"
                        elif key == "senate": label = "Senate District"
                        elif key == "congressional": label = "Congressional District"
                        
                        val = mappings.get(key)
                        mappings[key] = st.selectbox(
                            f"{label} Column",
                            ["None (Auto-detect / Unmapped)"] + voter_cols,
                            index=(voter_cols.index(val) + 1) if val in voter_cols else 0
                        )
                
                resolved_mappings = {}
                for k, v in mappings.items():
                    if v in ["None (Auto-detect / Unmapped)", "None (Turnout Dropoff Disabled)"]:
                        resolved_mappings[k] = None
                    else:
                        resolved_mappings[k] = v
                st.session_state["resolved_voter_mappings"] = resolved_mappings
    
    if st.button("Refresh Status", key="ref1"): st.rerun()

# --- TAB 2: CITY MAPPING MANAGER ---
with tab2:
    st.markdown("### City Assignment Manager (Optional)")
    if status['city']:
        st.success("✅ `srprec_city.csv` is loaded.")
        if st.button("🗑️ Remove City Mapping"): os.remove("data/srprec_city.csv"); st.rerun()
        
        try:
            city_cols = pd.read_csv("data/srprec_city.csv", nrows=1).columns.tolist()
        except:
            city_cols = []
            
        if city_cols:
            with st.expander("🗺️ City Mapping Column Mapping (Confirm / Override)", expanded=False):
                if "city_col_mappings" not in st.session_state:
                    c_cols_lower = {c.lower().strip(): c for c in city_cols}
                    s_def = c_cols_lower.get("srprec", city_cols[0] if city_cols else None)
                    c_def = c_cols_lower.get("city", city_cols[1] if len(city_cols) > 1 else city_cols[0])
                    st.session_state["city_col_mappings"] = {"srprec": s_def, "city": c_def}
                    
                c_maps = st.session_state["city_col_mappings"]
                c_maps["srprec"] = st.selectbox(
                    "SRPREC Column (GIS Precinct Key)",
                    city_cols,
                    index=city_cols.index(c_maps["srprec"]) if c_maps.get("srprec") in city_cols else 0,
                    key="city_map_srprec"
                )
                c_maps["city"] = st.selectbox(
                    "City Name Column",
                    city_cols,
                    index=city_cols.index(c_maps["city"]) if c_maps.get("city") in city_cols else 0,
                    key="city_map_city"
                )
                st.session_state["resolved_city_mappings"] = c_maps
    else:
        opt1, opt2, opt3 = st.tabs(["Upload CSV", "Generate via Shapefiles", "Manual Template"])
        with opt1:
            c_up = st.file_uploader("Upload pre-made City CSV", type=['csv'], key="c1")
            if c_up: save_uploaded(c_up, "srprec_city.csv"); st.rerun()
        with opt2:
            sc1, sc2 = st.columns(2)
            shp_srprec = sc1.file_uploader("SRPREC Shapes (.zip)", type=['zip'], key="csrd")
            shp_city = sc2.file_uploader("City Shapes (.zip)", type=['zip'], key="ccyd")
            if shp_srprec: save_uploaded(shp_srprec, "srprec_shapes.zip")
            if shp_city: save_uploaded(shp_city, "city_shapes.zip")
            if st.button("🗺️ Generate City Map & Extract Area"):
                if os.path.exists("data/srprec_shapes.zip"):
                    res_m = extract_precinct_metrics("data/srprec_shapes.zip", "data")
                if os.path.exists("data/srprec_shapes.zip") and os.path.exists("data/city_shapes.zip"):
                    res = generate_city_assignment_from_shapes("data/srprec_shapes.zip", "data/city_shapes.zip", "data")
                    if res["status"] == "success": st.success("Generated!"); st.rerun()

        with opt3:
            if st.button("Generate City Template"):
                res = generate_template(); st.success("OK")

# --- TAB 3: DISTRICT MAPPING MANAGER ---
with tab3:
    st.markdown("### Legislative District Manager (Optional)")
    if status['dist']:
        st.success("✅ `district_assignment.csv` is loaded.")
        if st.button("🗑️ Remove District Mapping"): os.remove("data/district_assignment.csv"); st.rerun()
        
        try:
            dist_cols = pd.read_csv("data/district_assignment.csv", nrows=1).columns.tolist()
        except:
            dist_cols = []
            
        if dist_cols:
            with st.expander("🗺️ District Mapping Column Mapping (Confirm / Override)", expanded=False):
                if "dist_col_mappings" not in st.session_state:
                    d_cols_lower = {c.lower().strip(): c for c in dist_cols}
                    s_def = d_cols_lower.get("srprec", dist_cols[0] if dist_cols else None)
                    a_def = d_cols_lower.get("assembly_district", next((c for c in dist_cols if 'assembly' in c.lower() or 'ad' in c.lower()), dist_cols[1] if len(dist_cols) > 1 else dist_cols[0]))
                    sup_def = d_cols_lower.get("supervisorial_district", next((c for c in dist_cols if 'supervisor' in c.lower() or 'sd' in c.lower()), dist_cols[2] if len(dist_cols) > 2 else dist_cols[0]))
                    st.session_state["dist_col_mappings"] = {"srprec": s_def, "assembly": a_def, "supervisorial": sup_def}
                    
                d_maps = st.session_state["dist_col_mappings"]
                d_maps["srprec"] = st.selectbox(
                    "SRPREC Column (GIS Precinct Key)",
                    dist_cols,
                    index=dist_cols.index(d_maps["srprec"]) if d_maps.get("srprec") in dist_cols else 0,
                    key="dist_map_srprec"
                )
                d_maps["assembly"] = st.selectbox(
                    "Assembly District Column",
                    dist_cols,
                    index=dist_cols.index(d_maps["assembly"]) if d_maps.get("assembly") in dist_cols else 0,
                    key="dist_map_assembly"
                )
                d_maps["supervisorial"] = st.selectbox(
                    "Supervisorial District Column",
                    dist_cols,
                    index=dist_cols.index(d_maps["supervisorial"]) if d_maps.get("supervisorial") in dist_cols else 0,
                    key="dist_map_supervisorial"
                )
                st.session_state["resolved_dist_mappings"] = d_maps
    elif status['is_mock_dist_present']:
        st.warning("⚠️ Mock `district_assignment.csv` is loaded. It will be ignored for production scoring.")
        if st.button("🗑️ Remove Mock District Mapping"): os.remove("data/district_assignment.csv"); st.rerun()
    else:
        opt1, opt2, opt3 = st.tabs(["Upload CSV", "Generate via Shapefiles", "Manual Template"])
        with opt1:
            d_up = st.file_uploader("Upload District CSV", type=['csv'], key="d1")
            if d_up: save_uploaded(d_up, "district_assignment.csv"); st.rerun()
        with opt2:
            sc1, sc2, sc3 = st.columns(3)
            shp_srprec2 = sc1.file_uploader("SRPREC Shapes (.zip)", type=['zip'], key="dscrd")
            shp_assem = sc2.file_uploader("Assembly Shapes (.zip)", type=['zip'], key="dasmd")
            shp_sup = sc3.file_uploader("Supervisor Shapes (.zip)", type=['zip'], key="dsupd")
            
            if shp_srprec2: save_uploaded(shp_srprec2, "srprec_shapes.zip")
            if shp_assem: save_uploaded(shp_assem, "assembly_shapes.zip")
            if shp_sup: save_uploaded(shp_sup, "supervisorial_shapes.zip")
            
            if st.button("🗺️ Generate District Map & Extract Area", type="primary"):
                if os.path.exists("data/srprec_shapes.zip"):
                    extract_precinct_metrics("data/srprec_shapes.zip", "data")
                if os.path.exists("data/srprec_shapes.zip") and os.path.exists("data/assembly_shapes.zip") and os.path.exists("data/supervisorial_shapes.zip"):
                    res = generate_district_assignment_from_shapes("data/srprec_shapes.zip", "data/assembly_shapes.zip", "data/supervisorial_shapes.zip", "data")
                    if res.get("status") == "success": st.success("Generated!"); st.rerun()

        with opt3:
            if st.button("Generate District Template"): generate_template()

# --- TAB 4: CONTEST DATA MANAGER ---
with tab4:
    st.markdown("### 📊 Contest Data Manager")
    st.info("Upload contest/results files. Classified contests are required for Final Production scoring.")
    
    active_contest_file = get_active_contest_file()
    
    c_up = st.file_uploader("Upload Contest/Results File (.csv, .tsv, .xlsx, .xls)", type=['csv', 'tsv', 'xlsx', 'xls'])
    if c_up:
        if st.session_state.get("last_processed_contest_upload") != c_up.name:
            for ext in ['.csv', '.tsv', '.xlsx', '.xls']:
                p = f"data/contest_data_input{ext}"
                if os.path.exists(p):
                    os.remove(p)
                    
            ext = os.path.splitext(c_up.name)[1].lower()
            new_path = f"data/contest_data_input{ext}"
            with open(new_path, "wb") as f:
                f.write(c_up.getbuffer())
            st.session_state["last_processed_contest_upload"] = c_up.name
            st.rerun()
        
    st.markdown("**OR Load directly from local absolute path:**")
    local_path = st.text_input("Paste absolute file path (e.g. D:\\Downloads\\November 4, 2025.xls)", key="contest_local_path")
    if st.button("Load from Local Path", key="btn_load_local_contest"):
        if local_path:
            clean_path = local_path.strip().strip('"').strip("'")
            if os.path.exists(clean_path):
                for ext in ['.csv', '.tsv', '.xlsx', '.xls']:
                    p = f"data/contest_data_input{ext}"
                    if os.path.exists(p):
                        os.remove(p)
                
                ext = os.path.splitext(clean_path)[1].lower()
                new_path = f"data/contest_data_input{ext}"
                try:
                    import shutil
                    shutil.copy(clean_path, new_path)
                    st.success(f"✅ Successfully loaded and copied `{os.path.basename(clean_path)}`!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Failed to copy file: {e}")
            else:
                st.error("File path does not exist. Please check the path and try again.")
        
    if active_contest_file:
        st.success(f"Loaded: `{os.path.basename(active_contest_file)}`")
        if st.button("🗑️ Remove Contest File"):
            if os.path.exists(active_contest_file):
                os.remove(active_contest_file)
            col_file = "outputs/contest_data_manager/contest_precinct_col.txt"
            if os.path.exists(col_file):
                os.remove(col_file)
            st.rerun()
            
        import contest_manager
        
        sheet_names = ["Default"]
        c_mtime = get_file_mtime(active_contest_file)
        res_load = cached_inspect_and_load_file(active_contest_file, mtime=c_mtime)
        if res_load["status"] == "error":
            if res_load.get("error_type") == "html_disguised_xls":
                st.error(res_load["message"])
            else:
                st.error(f"Failed to inspect file: {res_load['message']}")
        else:
            sheet_names = res_load["sheet_names"]
            selected_sheet = sheet_names[0]
            if len(sheet_names) > 1:
                selected_sheet = st.selectbox("Select Sheet/Table", sheet_names)
                
            inv = cached_generate_file_inventory(active_contest_file, sheet_name=selected_sheet, mtime=c_mtime)
            if inv["status"] == "success":
                st.markdown(f"**Rows:** {inv['row_count']} | **Columns:** {inv['col_count']}")
                st.dataframe(inv["df"].head(5))
                
                st.markdown("#### 🗺️ 1. Precinct Mapping Matcher")
                
                p_cols = inv["df"].columns.tolist()
                saved_prec_col = None
                col_file = "outputs/contest_data_manager/contest_precinct_col.txt"
                if os.path.exists(col_file):
                    with open(col_file, "r") as f:
                        saved_prec_col = f.read().strip()
                        
                default_idx = 0
                if saved_prec_col in p_cols:
                    default_idx = p_cols.index(saved_prec_col)
                elif inv["precinct_cols"]:
                    default_idx = p_cols.index(inv["precinct_cols"][0])
                    
                precinct_col = st.selectbox(
                    "Which column in this contest file identifies the precinct (PrecinctName)?",
                    p_cols,
                    index=default_idx
                )
                
                allow_low_coverage = False
                if precinct_col:
                    os.makedirs("outputs/contest_data_manager", exist_ok=True)
                    with open(col_file, "w") as f:
                        f.write(precinct_col)
                        
                    if status["voter"]:
                        v_mtime = get_file_mtime("data/voter_file.csv")
                        match_res = get_cached_precinct_match_report("data/voter_file.csv", v_mtime, active_contest_file, selected_sheet, c_mtime, precinct_col)
                        if match_res["status"] == "success":
                            rate = match_res["match_rate"]
                            st.markdown(f"- **Total Contest Precincts:** {match_res['total_contest_precincts']}")
                            st.markdown(f"- **Exact Match Count:** {match_res['exact_match_count']}")
                            st.markdown(f"- **Match Rate:** `{rate:.1f}%`")
                            
                            if rate < 80.0:
                                st.warning(f"⚠️ **Contest match rate is low ({rate:.1f}%).** Production Ranking is locked by default.")
                                allow_low_coverage = st.checkbox("Proceed with low-coverage contest data")
                            else:
                                st.success("✅ Precinct matching threshold passed.")
                                
                # Save override state in session state
                st.session_state["allow_low_coverage"] = allow_low_coverage

                st.markdown("#### 🗳️ 2. Contest Classification Wizard")
                config_list = contest_manager.load_classification_config()
                if config_list:
                    st.markdown("**Currently Classified Contests:**")
                    for idx, c in enumerate(config_list):
                        st.markdown(
                            f"- **{c['contest_name']}** ({c['contest_type']} in {c['year']}) -> "
                            f"Target: `{c['influence_component']}` | Weight: `{c['weight']}` "
                        )
                    
                    # Display scope status for first contest
                    c = config_list[0]
                    c_scope_type = c.get("scope_type", "unknown")
                    c_scope_val = c.get("scope_value", "")
                    c_scope_field = c.get("scope_field", "")
                    c_scope_conf = c.get("scope_confidence", "user_confirmed")
                    
                    st.markdown("##### 🔍 Active Contest Scope Status:")
                    curr_ad = st.session_state.get("target_ad_val", "ALL")
                    curr_sd = st.session_state.get("target_sd_val", "ALL")
                    curr_city = st.session_state.get("target_city_val", "ALL")
                    
                    active_universe = "Countywide"
                    if curr_sd != "ALL":
                        active_universe = f"Supervisorial District {curr_sd}"
                    elif curr_ad != "ALL":
                        active_universe = f"Assembly District {curr_ad}"
                    elif curr_city != "ALL":
                        active_universe = f"City {curr_city}"
                    
                    scope_match = "FAIL"
                    if c_scope_type == "countywide":
                        if active_universe == "Countywide":
                            scope_match = "PASS"
                        else:
                            scope_match = "PASS (Broader)"
                    elif c_scope_type == "supervisorial_district":
                        if active_universe == f"Supervisorial District {c_scope_val}":
                            scope_match = "PASS"
                    elif c_scope_type == "assembly_district":
                        if active_universe == f"Assembly District {c_scope_val}":
                            scope_match = "PASS"
                    elif c_scope_type == "city":
                        if active_universe == f"City {c_scope_val}":
                            scope_match = "PASS"
                            
                    scope_details_lbl = "Unknown"
                    if c_scope_type == "countywide":
                        scope_details_lbl = "Countywide"
                    elif c_scope_type != "unknown":
                        scope_details_lbl = f"{c_scope_type.replace('_', ' ').title()} {c_scope_val}"
                        
                    st.write(f"- **Contest Scope:** {scope_details_lbl}")
                    st.write(f"- **Scope source:** {c_scope_conf.replace('_', ' ')}")
                    st.write(f"- **Selected Universe:** {active_universe}")
                    if "PASS" in scope_match:
                        st.success(f"🟢 **Scope Match:** {scope_match}")
                    else:
                        st.error("🔴 **Scope Match:** FAIL")
                        st.warning("⚠️ Production ranking is blocked. Adjust the target universe filters in the 'Run & Results' tab to match the contest scope.")
                        
                    if st.button("🗑️ Clear All Classifications"):
                        contest_manager.save_classification_config([])
                        st.rerun()
                        
                st.write("---")
                st.write("**Add Contest/Result Column:**")
                
                c_name = st.text_input("Contest Name (e.g. 2024 Presidential)", key="wizard_contest_name")
                
                # Scope auto-detection logic based on name
                suggested_scope_type = "unknown"
                suggested_scope_field = ""
                suggested_scope_value = ""
                suggestion_msg = None
                
                if c_name:
                    import re
                    # Match D4, District 4, Supervisor D4, Supervisorial District 4, etc.
                    sd_m = re.search(r'(?:supervisorial\s+district|supervisor|sd|d|district)\s*([1-9])\b', c_name, re.IGNORECASE)
                    if sd_m:
                        val = sd_m.group(1)
                        suggested_scope_type = "supervisorial_district"
                        suggested_scope_field = "Supervisorial_District"
                        suggested_scope_value = val
                        suggestion_msg = f"Detected possible contest scope: Supervisorial District {val}."
                    else:
                        if re.search(r'(?:countywide|measure|proposition|prop)\b', c_name, re.IGNORECASE):
                            suggested_scope_type = "countywide"
                            suggested_scope_field = "County"
                            suggested_scope_value = "Sonoma"
                            suggestion_msg = "Detected possible contest scope: Countywide."
                
                if suggestion_msg:
                    st.info(f"💡 **{suggestion_msg}** Please confirm or adjust the scope settings below.")
                
                c_year = st.number_input("Year", min_value=2000, max_value=2030, value=2024, key="wizard_year")
                c_elec = st.selectbox("Election Type", ["General", "Primary", "Special", "Local", "Other"], key="wizard_election_type")
                c_type = st.selectbox(
                    "Contest Type", 
                    ["Candidate", "Initiative / ballot measure", "Turnout", "Party baseline", "Other"],
                    key="wizard_contest_type"
                )
                
                influence_target = st.selectbox(
                    "Influence Target component",
                    ["Support Score", "Persuasion Score", "Turnout Score", "Issue Alignment Score", "Confidence Only"],
                    key="wizard_influence_target"
                )
                c_weight = st.slider("Weight", 0.0, 1.0, 0.5, key="wizard_weight")
                
                # --- SCOPE METADATA FIELDS ---
                scope_options = [
                    "Countywide", "Supervisorial District", "Assembly District", "Senate District", 
                    "Congressional District", "City", "City Council District", "School District", 
                    "Special District", "Custom Precinct Set", "Unknown / Not Sure"
                ]
                scope_keys = {
                    "Countywide": "countywide",
                    "Supervisorial District": "supervisorial_district",
                    "Assembly District": "assembly_district",
                    "Senate District": "senate_district",
                    "Congressional District": "congressional_district",
                    "City": "city",
                    "City Council District": "city_council_district",
                    "School District": "school_district",
                    "Special District": "special_district",
                    "Custom Precinct Set": "custom_precinct_set",
                    "Unknown / Not Sure": "unknown"
                }
                
                # Pre-select index if suggested
                default_scope_idx = len(scope_options) - 1 # Default to Unknown
                if suggested_scope_type in scope_keys.values():
                    for k, v in scope_keys.items():
                        if v == suggested_scope_type:
                            default_scope_idx = scope_options.index(k)
                            break
                            
                scope_label = st.selectbox("What geographic universe does this contest cover?", scope_options, index=default_scope_idx, key="wizard_scope_label")
                scope_type = scope_keys[scope_label]
                
                # Map fields
                field_mapping = {
                    "countywide": "County",
                    "supervisorial_district": "Supervisorial_District",
                    "assembly_district": "Assembly_District",
                    "senate_district": "Senate_District",
                    "congressional_district": "Congressional_District",
                    "city": "CITY",
                    "city_council_district": "City_Council_District",
                    "school_district": "School_District",
                    "special_district": "Special_District",
                    "custom_precinct_set": "PrecinctName",
                    "unknown": ""
                }
                scope_field = field_mapping[scope_type]
                
                scope_value = ""
                
                ad_opts, sd_opts, city_opts = [], [], []
                if status["voter"]:
                    try:
                        ad_opts, sd_opts, city_opts, _ = get_cached_geo_options(
                            "data/voter_file.csv", 
                            "data/mprec_srprec.csv", 
                            "data/srprec_city.csv", 
                            "data/district_assignment.csv",
                            False, None, None, None, None
                        )
                    except:
                        pass
                
                if scope_type == "countywide":
                    scope_value = st.text_input("County Name", value=suggested_scope_value or "Sonoma", key="wizard_scope_value_county")
                elif scope_type == "supervisorial_district":
                    opts = [x for x in sd_opts if x != "ALL"] if sd_opts else ["1", "2", "3", "4", "5"]
                    def_val_idx = 0
                    if suggested_scope_value in opts:
                        def_val_idx = opts.index(suggested_scope_value)
                    scope_value = st.selectbox("Which Supervisorial District?", opts, index=def_val_idx, key="wizard_scope_value_sd")
                elif scope_type == "assembly_district":
                    opts = [x for x in ad_opts if x != "ALL"] if ad_opts else []
                    scope_value = st.selectbox("Which Assembly District?", opts, key="wizard_scope_value_ad")
                elif scope_type == "city":
                    opts = [x for x in city_opts if x != "ALL"] if city_opts else []
                    scope_value = st.selectbox("Which City?", opts, key="wizard_scope_value_city")
                elif scope_type != "unknown":
                    scope_value = st.text_input("District / Scope Value", value=suggested_scope_value, key="wizard_scope_value_manual")
                
                num_cols = inv["df"].columns.tolist()
                
                fav_col = st.selectbox("Favorable Votes / Yes / Dem Column", num_cols, key="wizard_fav_col")
                opp_col = st.selectbox("Opposition / No Column (for Candidate type)", ["None"] + num_cols, key="wizard_opp_col")
                tot_col = st.selectbox("Total votes column (for Initiative/Baseline type)", ["None"] + num_cols, key="wizard_tot_col")
                reg_col = st.selectbox("Registered Voters column (for Turnout type)", ["None"] + num_cols, key="wizard_reg_col")
                
                c_cross = st.checkbox("Uses Official Sonoma Crosswalk", value=False, key="wizard_uses_crosswalk")
                c_reg_pdf_val = ""
                c_voting_pdf_val = ""
                if c_cross:
                    metadata_fm = file_manager.load_file_metadata()
                    pdf_files = [f for f in metadata_fm.keys() if f.lower().endswith(".pdf")]
                    c_reg_pdf = st.selectbox("Select Regular-to-Voting Crosswalk PDF (e.g. ewmr010)", ["None"] + pdf_files, key="wizard_reg_pdf")
                    c_voting_pdf = st.selectbox("Select Voting-to-Regular Crosswalk PDF (e.g. ewmr008)", ["None"] + pdf_files, key="wizard_voting_pdf")
                    c_reg_pdf_val = "" if c_reg_pdf == "None" else os.path.join("data", c_reg_pdf)
                    c_voting_pdf_val = "" if c_voting_pdf == "None" else os.path.join("data", c_voting_pdf)
                
                if st.button("➕ Add Contest to Scoring Model", key="wizard_submit"):
                    if c_name:
                        c_def = {
                            "contest_name": c_name,
                            "name": c_name,
                            "year": int(c_year),
                            "election_type": c_elec,
                            "contest_type": c_type,
                            "influence_component": influence_target,
                            "weight": float(c_weight),
                            "favorable_col": fav_col,
                            "scope_type": scope_type,
                            "scope_field": scope_field,
                            "scope_value": str(scope_value),
                            "scope_confidence": "user_confirmed",
                            "scope_source": "contest_classification_wizard",
                            "scope_user_confirmed": True,
                            "uses_official_crosswalk": c_cross,
                            "crosswalk_reg_to_voting_file": c_reg_pdf_val,
                            "crosswalk_voting_to_reg_file": c_voting_pdf_val
                        }
                        if c_type == "Candidate":
                            c_def["opposition_col"] = opp_col
                        elif c_type == "Initiative / ballot measure":
                            c_def["total_col"] = tot_col
                        elif c_type == "Turnout":
                            c_def["ballots_col"] = fav_col
                            c_def["reg_col"] = reg_col
                        elif c_type == "Party baseline":
                            c_def["total_col"] = tot_col
                            
                        config_list.append(c_def)
                        contest_manager.save_classification_config(config_list)
                        st.success(f"Successfully added contest '{c_name}'!")
                        st.rerun()

# --- TAB 5: RUN & RESULTS ---
with tab5:
    st.markdown("### Target Extraction Parameters")
    
    if not status['voter']:
        st.error("Missing Core Dependency: Voter File is required. Scoring Pipeline Locked.")
    else:
        # Determine Mode
        is_override = st.session_state.get("allow_low_coverage", False)
        
        # Determine options dynamically
        ad_opts, sd_opts, city_opts, geo_metadata = get_cached_geo_options(
            "data/voter_file.csv", 
            "data/mprec_srprec.csv", 
            "data/srprec_city.csv", 
            "data/district_assignment.csv",
            derive_sonoma_sd,
            st.session_state.get("resolved_voter_mappings"),
            st.session_state.get("resolved_mprec_mappings"),
            st.session_state.get("resolved_city_mappings"),
            st.session_state.get("resolved_dist_mappings")
        )
        
        st.markdown("#### 🗺️ Geography Sources Engaged:")
        m_cols = st.columns(4)
        m_cols[0].markdown(f"**Supervisorial District:**\n\nSource: `{geo_metadata['supervisorial']['source']}`\n\nConfidence: `{geo_metadata['supervisorial']['confidence']}`")
        m_cols[1].markdown(f"**Assembly District:**\n\nSource: `{geo_metadata['assembly']['source']}`\n\nConfidence: `{geo_metadata['assembly']['confidence']}`")
        m_cols[2].markdown(f"**Senate District:**\n\nSource: `{geo_metadata['senate']['source']}`\n\nConfidence: `{geo_metadata['senate']['confidence']}`")
        m_cols[3].markdown(f"**City:**\n\nSource: `{geo_metadata['city']['source']}`\n\nConfidence: `{geo_metadata['city']['confidence']}`")
        
        # Ensure session state variables exist
        if "target_ad_val" not in st.session_state:
            st.session_state["target_ad_val"] = "ALL"
        if "target_sd_val" not in st.session_state:
            st.session_state["target_sd_val"] = "ALL"
        if "target_city_val" not in st.session_state:
            st.session_state["target_city_val"] = "ALL"

        st.info("Select Target Region constraints. Filters are applied dynamically.")
        c1, c2, c3 = st.columns(3)
        
        # Calculate indices
        idx_ad = ad_opts.index(st.session_state["target_ad_val"]) if st.session_state["target_ad_val"] in ad_opts else 0
        idx_sd = sd_opts.index(st.session_state["target_sd_val"]) if st.session_state["target_sd_val"] in sd_opts else 0
        idx_city = city_opts.index(st.session_state["target_city_val"]) if st.session_state["target_city_val"] in city_opts else 0
        
        target_ad = c1.selectbox("Filter by Assembly District", ad_opts, index=idx_ad)
        target_sd = c2.selectbox("Filter by Supervisorial District", sd_opts, index=idx_sd)
        target_city = c3.selectbox("Filter by City", city_opts, index=idx_city)
        
        # Keep session state in sync
        st.session_state["target_ad_val"] = target_ad
        st.session_state["target_sd_val"] = target_sd
        st.session_state["target_city_val"] = target_city
        
        target_params = {
            "ad": None if target_ad == "ALL" else target_ad,
            "sd": None if target_sd == "ALL" else target_sd,
            "city": None if target_city == "ALL" else target_city,
        }

        # Determine Contest Scope and Universe Relationship
        relationship = "exact_match"
        scope_details = "None"
        scope_field = ""
        scope_value = ""
        scope_type = "unknown"
        scope_confirmed = "No"
        
        # Load classifications
        import contest_manager
        configs = contest_manager.load_classification_config()
        first_contest = configs[0] if configs else None
        
        if production_ranking_ready and first_contest:
            scope_type = first_contest.get("scope_type", "unknown")
            scope_field = first_contest.get("scope_field", "")
            scope_value = first_contest.get("scope_value", "")
            scope_source = first_contest.get("scope_source", "none")
            c_name = first_contest.get("contest_name", first_contest.get("name", ""))
            scope_confirmed = "Yes" if first_contest.get("scope_user_confirmed") else "No"
            
            if scope_type == "unknown":
                relationship = "unknown_scope"
                scope_details = "Unknown scope"
            elif scope_type == "countywide":
                scope_details = "Countywide"
                if not target_params["ad"] and not target_params["sd"] and not target_params["city"]:
                    relationship = "exact_match"
                else:
                    relationship = "contest_broader_than_selected_universe"
            else:
                scope_details = f"{scope_type.replace('_', ' ').title()} {scope_value}"
                
                # Check for field availability
                field_avail = True
                if scope_type == "supervisorial_district" and (not sd_opts or len(sd_opts) <= 1):
                    field_avail = False
                elif scope_type == "assembly_district" and (not ad_opts or len(ad_opts) <= 1):
                    field_avail = False
                elif scope_type == "city" and (not city_opts or len(city_opts) <= 1):
                    field_avail = False
                    
                if not field_avail:
                    relationship = "scope_field_unavailable"
                else:
                    # Check for value existence
                    val_exists = True
                    if scope_type == "supervisorial_district" and str(scope_value) not in sd_opts:
                        val_exists = False
                    elif scope_type == "assembly_district" and str(scope_value) not in ad_opts:
                        val_exists = False
                    elif scope_type == "city" and str(scope_value) not in city_opts:
                        val_exists = False
                        
                    if not val_exists:
                        relationship = "scope_value_not_found"
                    else:
                        # Standard relationship check
                        target_key = None
                        if scope_type == "supervisorial_district":
                            target_key = "sd"
                        elif scope_type == "assembly_district":
                            target_key = "ad"
                        elif scope_type == "city":
                            target_key = "city"
                            
                        active_filters = {k: v for k, v in target_params.items() if v is not None}
                        
                        if target_key is not None:
                            val = target_params.get(target_key)
                            if val is not None and str(val).strip() == str(scope_value).strip():
                                other_filters = {k: v for k, v in active_filters.items() if k != target_key}
                                if not other_filters:
                                    relationship = "exact_match"
                                else:
                                    relationship = "contest_broader_than_selected_universe"
                            else:
                                relationship = "contest_narrower_than_selected_universe"
                        else:
                            relationship = "contest_narrower_than_selected_universe"

        # Render warning panels & checkboxes
        run_btn_label = "🚀 Execute Production Precinct Scoring"
        btn_disabled = False
        block_execution = False
        warning_msg = ""
        
        # Provenance and Mock detection warning
        uses_mock = False
        if active_contest_file:
            acf_lower = active_contest_file.replace("\\", "/").lower()
            if "tests/" in acf_lower or "fixtures/" in acf_lower or "mock" in acf_lower or ("test" in acf_lower and "contest" not in acf_lower):
                uses_mock = True

            elif os.path.exists(active_contest_file) and os.path.getsize(active_contest_file) < 15000:
                try:
                    df_c = pd.read_csv(active_contest_file, nrows=3)
                    if "harris_dem" in [c.lower() for c in df_c.columns]:
                        uses_mock = True
                except:
                    pass
        if uses_mock:
            st.warning("⚠️ **Mock/test contest file detected.** Production ranking is blocked. This file may be used only for TEST_MODE validation.")
            run_btn_label = "🔒 Execute Scoring (Blocked by Mock File)"
            btn_disabled = True
            block_execution = True

        # Legacy scope warning
        if production_ranking_ready and first_contest:
            is_legacy = (scope_source == "legacy")
            if is_legacy:
                st.warning("⚠️ Contest scope is based on a legacy/default value and has not been confirmed. Production ranking blocked until scope is confirmed.")
                run_btn_label = "🔒 Execute Scoring (Blocked by Unconfirmed Legacy Scope)"
                btn_disabled = True
                block_execution = True

            # Configuration truth conflict warning
            has_district_name = any(x in c_name.upper() for x in ["D4", "DISTRICT 4", "SUPERVISOR D4", "SUPERVISORIAL DISTRICT 4"])
            is_countywide_or_legacy = scope_type in ["countywide", "unknown", "legacy"]
            if has_district_name and is_countywide_or_legacy:
                st.error("⚠️ **Configuration Truth Conflict:** The contest name implies a district-specific scope, but it is configured as countywide/legacy. Production ranking is blocked unless confirmed.")
                scope_override_confirmed = st.checkbox("I confirm this contest is countywide despite the district-specific name.", key="scope_override_confirmed")
                if not scope_override_confirmed:
                    run_btn_label = "🔒 Execute Scoring (Blocked by Scope Conflict)"
                    btn_disabled = True
                    block_execution = True
        
        if production_ranking_ready and first_contest:
            if relationship == "unknown_scope":
                st.error("🔒 **Contest scope is unknown.** Production ranking requires a confirmed contest scope. Please classify the scope in the 'Contest Data Manager' tab.")
                run_btn_label = "🔒 Execute Scoring (Locked)"
                btn_disabled = True
                block_execution = True
            elif relationship == "scope_field_unavailable":
                st.error(f"❌ Cannot apply {scope_type.replace('_', ' ').title()} scope because the field is unavailable or unvalidated. Map a supervisorial/district column or enable Sonoma precinct-prefix validation.")
                run_btn_label = "🔒 Execute Scoring (Locked)"
                btn_disabled = True
                block_execution = True
            elif relationship == "scope_value_not_found":
                st.error(f"❌ Mapped scope value '{scope_value}' not found in the voter file. Please map the correct scope.")
                run_btn_label = "🔒 Execute Scoring (Locked)"
                btn_disabled = True
                block_execution = True
            elif relationship == "contest_narrower_than_selected_universe":
                active_universe_label = "Countywide"
                active_filters = [f"{k.upper()}: {v}" for k, v in target_params.items() if v is not None]
                if active_filters:
                    active_universe_label = ", ".join(active_filters)
                    
                warning_msg = f"Production ranking blocked: this contest appears to cover {scope_details}, but the selected universe is {active_universe_label}. Select {scope_details} or mark the contest as countywide if that is correct."
                st.error(f"🔒 {warning_msg}")
                
                # Buttons
                c_btn1, c_btn2 = st.columns(2)
                if c_btn1.button("Apply contest scope to selected universe", key="btn_apply_scope"):
                    if scope_type == "supervisorial_district":
                        st.session_state["target_sd_val"] = str(scope_value)
                        st.session_state["target_ad_val"] = "ALL"
                        st.session_state["target_city_val"] = "ALL"
                    elif scope_type == "assembly_district":
                        st.session_state["target_ad_val"] = str(scope_value)
                        st.session_state["target_sd_val"] = "ALL"
                        st.session_state["target_city_val"] = "ALL"
                    elif scope_type == "city":
                        st.session_state["target_city_val"] = str(scope_value)
                        st.session_state["target_ad_val"] = "ALL"
                        st.session_state["target_sd_val"] = "ALL"
                    st.session_state["contest_scope_auto_applied"] = True
                    st.rerun()
                    
                c_btn2.markdown("👈 **Go to the 'Contest Data Manager' tab above to edit contest scope.**")
                
                override_scope = st.checkbox("Override scope mismatch (Caution: Limits verdict to PRODUCTION_READY_WITH_CAUTION)", key="override_scope_mismatch")
                if not override_scope:
                    run_btn_label = "🔒 Execute Scoring (Blocked by Scope Mismatch)"
                    btn_disabled = True
                    block_execution = True
            elif relationship == "contest_broader_than_selected_universe":
                st.warning(f"⚠️ **Note:** The selected universe is narrower than the contest scope ({scope_details}). This is allowed if coverage passes.")
                st.session_state["override_scope_mismatch"] = False
                
        # Build state displays
        matches_scope_label = "Yes" if relationship in ["exact_match", "contest_broader_than_selected_universe"] else "No"
        if not production_ranking_ready:
            matches_scope_label = "N/A"
            
        last_coverage_str = "Run pipeline to calculate"
        last_verdict_str = "N/A"
        try:
            log_path = "outputs/final_validation/active_overrides_log.json"
            if os.path.exists(log_path):
                import json
                with open(log_path, "r") as f:
                    log_data = json.load(f)
                    log_filters = log_data.get("selected_universe_filters", {})
                    curr_filters = {k: v for k, v in target_params.items() if v is not None}
                    mapped_curr = {}
                    if curr_filters.get("sd"): mapped_curr["Supervisorial_District"] = curr_filters["sd"]
                    if curr_filters.get("ad"): mapped_curr["Assembly_District"] = curr_filters["ad"]
                    if curr_filters.get("city"): mapped_curr["CITY"] = curr_filters["city"]
                    
                    if mapped_curr == log_filters:
                        last_coverage_str = f"{log_data.get('row_level_coverage', 0.0):.2f}%"
                        last_verdict_str = log_data.get("production_readiness_verdict", "N/A")
        except:
            pass
            
        st.markdown("### Production Ranking Status")
        st.write(f"- **Voter file:** {'loaded' if status['voter'] else 'missing'}")
        st.write(f"- **Contest data:** {'classified' if production_ranking_ready else 'not loaded'}")
        st.write(f"- **Contest scope:** {'confirmed' if scope_confirmed == 'Yes' else 'unconfirmed'} ({scope_details})")
        st.write(f"- **Selected universe matches contest scope:** {matches_scope_label}")
        st.write(f"- **Selected-universe contest coverage:** {last_coverage_str}")
        st.write(f"- **Production readiness verdict:** {last_verdict_str}")
        
        if last_verdict_str in ["CONTEST_DATA_INCOMPLETE_FOR_SELECTED_UNIVERSE", "LIMITED_CONTEST_COVERAGE_PREVIEW"]:
            st.warning("⚠️ **The selected contest and universe match, but the uploaded Statement of Votes file only covers 55 of 268 D4 precincts, or 20.52%. This appears to be an incomplete SOV file, not a precinct-format problem. Upload the complete D4 Statement of Votes file before using production rankings.**")
            st.markdown("""
##### 📋 Complete SOV File Intake Checklist
* **Complete Sonoma County Statement of Votes** for the relevant election
* **Must include all precinct rows** for Supervisorial District 4
* **Must include Precinct column**
* **Must include MELANIE BAGBY - Total Votes**
* **Must include TOM SCHWEDHELM - Total Votes**
""")
        st.write("---")

        if st.button(run_btn_label, type="primary", use_container_width=True, disabled=btn_disabled):

            with st.spinner("Running Realigned Precinct Algorithms..."):
                if "latest_pipeline_result" in st.session_state:
                    del st.session_state["latest_pipeline_result"]
                if "latest_output_manifest" in st.session_state:
                    del st.session_state["latest_output_manifest"]
                if "latest_run_context" in st.session_state:
                    del st.session_state["latest_run_context"]

                if not status['metrics'] and status['shp_srprec']: extract_precinct_metrics("data/srprec_shapes.zip", "data")
                if not status['city'] and status['city_shapefiles_ready']: generate_city_assignment_from_shapes("data/srprec_shapes.zip", "data/city_shapes.zip", "data")
                if not status['dist'] and status['dist_shapefiles_ready']: generate_district_assignment_from_shapes("data/srprec_shapes.zip", "data/assembly_shapes.zip", "data/supervisorial_shapes.zip", "data")

                result = run_pipeline(
                    weights=weights, 
                    target_params=target_params, 
                    allow_mock=False,
                    derive_sonoma_sd=derive_sonoma_sd,
                    contest_file_path=active_contest_file,
                    contest_prec_col=contest_prec_col,
                    contest_influence_weight=contest_influence_weight,
                    allow_low_coverage_contest=is_override,
                    voter_col_mappings=st.session_state.get("resolved_voter_mappings"),
                    mprec_col_mappings=st.session_state.get("resolved_mprec_mappings"),
                    city_col_mappings=st.session_state.get("resolved_city_mappings"),
                    dist_col_mappings=st.session_state.get("resolved_dist_mappings"),
                    election_context=election_context,
                    target_turnout_override=target_turnout_override,
                    enforce_size_guardrail=enforce_size_guardrail,
                    override_scope_mismatch=st.session_state.get("override_scope_mismatch", False),
                    contest_scope_auto_applied=st.session_state.get("contest_scope_auto_applied", False),
                    run_mode="USER_DASHBOARD_MODE",
                    trigger_source="streamlit_ui",
                    scope_override_confirmed=st.session_state.get("scope_override_confirmed", False)
                )
                
            if result.get("status") == "validation_error":
                st.error("❌ Pipeline explicitly blocked by data validation failure:")
                st.error(result["message"])
                for w in result.get("warnings", []): st.warning(f"⚠ {w}")
                
            elif result.get("status") == "success":
                st.session_state["latest_pipeline_result"] = result
                st.session_state["latest_run_context"] = {
                    "run_mode": "USER_DASHBOARD_MODE",
                    "active_voter_file": "data/voter_file.csv",
                    "active_contest_file": active_contest_file or "",
                    "active_cross_reference_files": [
                        r"D:\Downloads\ewmr010_regabsvotpctxref_2026-06-02.pdf",
                        r"D:\Downloads\ewmr008_votabsregpctxref_2026-06-02.pdf"
                    ] if os.path.exists(r"D:\Downloads\ewmr010_regabsvotpctxref_2026-06-02.pdf") else [],
                    "readiness_verdict": result.get("verdict", "unknown")
                }
                st.session_state["latest_output_manifest"] = build_output_manifest(st.session_state["latest_run_context"])
                st.rerun()
            else:
                st.error("❌ Critical Pipeline Crash.")
                st.code(result.get("error", "Unknown Fault"))

        # Render persistent results from session state
        if st.session_state.get("latest_pipeline_result") and st.session_state.get("latest_output_manifest"):
            result = st.session_state["latest_pipeline_result"]
            manifest = st.session_state["latest_output_manifest"]
            run_context = st.session_state.get("latest_run_context")
            top_df = result.get("top_precincts", pd.DataFrame())
            
            if top_df.empty:
                st.error("❌ THE PIPELINE EXECUTED, BUT 0 PRECINCTS MATCHED YOUR TARGETING SELECTION.")
            else:
                st.success("✅ Strict Data Execution Complete.")
                
                if "Contest_Confidence" in top_df.columns and not top_df.empty:
                    avg_conf = top_df["Contest_Confidence"].mean() * 100.0
                    if active_contest_file and contest_prec_col and avg_conf < 100.0:
                        st.warning(f"⚠️ **Warning:** Contest data covers only {avg_conf:.1f}% of precincts.")

                try:
                    voter_count = len(pd.read_csv("data/voter_file.csv", usecols=[0]))
                except:
                    voter_count = 0

                m1, m2, m3 = st.columns(3)
                m1.markdown(f'<div class="metric-box"><div class="metric-title">Voters</div><div class="metric-value">{voter_count:,}</div></div>', unsafe_allow_html=True)
                m2.markdown(f'<div class="metric-box"><div class="metric-title">Valid Precincts Scored</div><div class="metric-value">{len(top_df):,}</div></div>', unsafe_allow_html=True)
                m3.markdown(f'<div class="metric-box"><div class="metric-title">Mode</div><div class="metric-value" style="color: {"green" if result.get("has_contest") else "orange"};">{"Production" if result.get("has_contest") else "Base Preview"}</div></div>', unsafe_allow_html=True)
                
                if "Rank_Change" in top_df.columns and result.get("has_contest"):
                    changed = top_df[top_df["Rank_Change"] != 0]
                    if not changed.empty:
                        st.markdown("#### 🔄 Rank Shift Analysis (Base Rank vs Final Rank)")
                        st.write(f"- **Precincts with Rank Changes:** {len(changed)} out of {len(top_df)} precincts")
                        st.write(f"- **Average Rank Shift:** {changed['Rank_Change'].abs().mean():.1f} positions")
                
                render_download_panel(manifest, run_context)
                
                st.markdown("---")
                st.markdown("## 🔍 Proof Exports for Review")
                
                st.markdown("### 📊 Key Status Summary")
                ks1, ks2, ks3 = st.columns(3)
                
                verdict_style = "color: green;" if result.get("verdict") in ["PRODUCTION_READY", "PRODUCTION_READY_WITH_INHERITED_CONTEST_SIGNALS"] else "color: orange;" if result.get("verdict") in ["PRODUCTION_READY_WITH_CAUTION", "LIMITED_CONTEST_COVERAGE_PREVIEW"] else "color: red;"
                ks1.markdown(f"**Production readiness verdict:** <span style='{verdict_style} font-weight: bold;'>{result.get('verdict')}</span>", unsafe_allow_html=True)
                ks1.write(f"**Total precincts in selected universe:** {result.get('total_precincts')}")
                ks1.write(f"**Matched precincts:** {result.get('matched_precincts')}")
                ks1.write(f"**Unmatched precincts:** {result.get('unmatched_precincts')}")
                
                ks2.write(f"**Countywide contest coverage:** {result.get('countywide_coverage', 0.0):.2f}%")
                ks2.write(f"**Selected-universe contest coverage:** {result.get('universe_coverage', 0.0):.2f}%")
                ks2.write(f"**Contest influence weight:** {result.get('contest_influence_weight', 0.0):.2f}")
                
                ks3.write(f"**Top 50 precincts without contest match:** {result.get('top_50_unmatched')}")
                ks3.write(f"**Tiny precincts in top 50:** {result.get('tiny_in_top_50')}")
                ks3.write(f"**Active normalization rules:** {result.get('active_normalization_rules')}")
                ks3.write(f"**Active overrides:** {result.get('active_overrides')}")
                
                st.markdown("### 🏆 Production Priority Preview (First 100 Rows)")
                prod_csv_path = "outputs/final_rankings/production_priority_precincts.csv"
                if os.path.exists(prod_csv_path):
                    prod_preview_df = pd.read_csv(prod_csv_path)
                    prod_abs = os.path.abspath(prod_csv_path)
                    st.write(f"📁 **Local Path:** `{prod_abs}` | 📊 **Row Count:** `{len(prod_preview_df)}` rows")
                    st.dataframe(prod_preview_df.head(100), use_container_width=True)
                else:
                    st.warning("production_priority_precincts.csv not found.")
                    
                st.markdown("### 📝 Precinct Normalization Audit Preview (First 250 Rows)")
                audit_csv_path = "outputs/contest_data_manager/precinct_normalization_audit.csv"
                if os.path.exists(audit_csv_path):
                    audit_preview_df = pd.read_csv(audit_csv_path)
                    audit_abs = os.path.abspath(audit_csv_path)
                    st.write(f"📁 **Local Path:** `{audit_abs}` | 📊 **Row Count:** `{len(audit_preview_df)}` rows")
                    st.dataframe(audit_preview_df.head(250), use_container_width=True)
                else:
                    st.warning("precinct_normalization_audit.csv not found.")

                st.markdown("### 🗺️ Precinct Crosswalk Outputs")
                cw_summary_path = "outputs/precinct_crosswalk/crosswalk_validation_summary.md"
                if os.path.exists(cw_summary_path):
                    with open(cw_summary_path, "r", encoding="utf-8") as f:
                        st.markdown(f.read())

# --- CONTEST SIGNAL MANAGER TAB ---
with tab6:
    st.markdown("### 📈 6. Contest Signal Manager")
    st.markdown("Build, validate, and preview multi-contest signal models for precinct-level targeting.")
    
    # 1. Campaign Profile Panel
    st.subheader("👤 Current Campaign Profile")
    
    profile_path = "outputs/contest_signal_model/current_campaign_profile.json"
    campaign_profile = {}
    if os.path.exists(profile_path):
        try:
            with open(profile_path, "r", encoding="utf-8") as f:
                campaign_profile = json.load(f)
        except:
            pass
            
    campaign_name = st.text_input("Campaign Name", value=campaign_profile.get("campaign_name", "My Campaign"))
    current_contest_name = st.text_input("Current Contest Name", value=campaign_profile.get("current_contest_name", "My Contest"))
    supported_side = st.text_input("Supported Candidate/Cause/Side", value=campaign_profile.get("supported_side", ""))
    opposed_side = st.text_input("Opposed Candidate/Cause/Side (Optional)", value=campaign_profile.get("opposed_side", ""))
    
    contest_type_opts = [
        "Candidate race", "Ballot measure", "Turnout result", "Party baseline", "Issue alignment",
        "Local candidate performance", "Statewide candidate performance", "Federal candidate performance",
        "Environmental issue contest", "Labor issue contest", "Housing issue contest",
        "Tax/revenue issue contest", "Public safety issue contest", "Custom"
    ]
    def_type_idx = contest_type_opts.index(campaign_profile.get("contest_type")) if campaign_profile.get("contest_type") in contest_type_opts else 0
    contest_type = st.selectbox("Contest Type", contest_type_opts, index=def_type_idx)
    
    election_date = st.text_input("Election Date", value=campaign_profile.get("election_date", ""))
    selected_geography = st.text_input("Selected Geography Name", value=campaign_profile.get("selected_geography", ""))
    
    scope_type_opts = ["countywide", "supervisorial_district", "assembly_district", "city", "other"]
    def_scope_idx = scope_type_opts.index(campaign_profile.get("selected_scope_type")) if campaign_profile.get("selected_scope_type") in scope_type_opts else 0
    selected_scope_type = st.selectbox("Scope Type", scope_type_opts, index=def_scope_idx)
    selected_scope_value = st.text_input("Scope Value (e.g. 4 for SD4)", value=campaign_profile.get("selected_scope_value", ""))
    
    goal_opts = ["Elect candidate", "Defeat candidate", "Pass ballot measure", "Defeat ballot measure", "Turnout operation", "Persuasion operation", "Research / exploratory"]
    def_goal_idx = goal_opts.index(campaign_profile.get("primary_campaign_goal")) if campaign_profile.get("primary_campaign_goal") in goal_opts else 0
    primary_campaign_goal = st.selectbox("Primary Campaign Goal", goal_opts, index=def_goal_idx)
    
    if st.button("💾 Save Campaign Profile"):
        os.makedirs("outputs/contest_signal_model", exist_ok=True)
        campaign_profile = {
            "campaign_name": campaign_name,
            "current_contest_name": current_contest_name,
            "supported_side": supported_side,
            "opposed_side": opposed_side,
            "contest_type": contest_type,
            "election_date": election_date,
            "selected_geography": selected_geography,
            "selected_scope_type": selected_scope_type,
            "selected_scope_value": selected_scope_value,
            "primary_campaign_goal": primary_campaign_goal,
            "created_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        with open(profile_path, "w", encoding="utf-8") as f:
            json.dump(campaign_profile, f, indent=2)
        st.success("Campaign Profile Saved!")
        st.rerun()
        
    st.markdown("---")
    
    # 2. Contest Library Selection & Management Panel
    st.subheader("📚 Contest Library")
    library_path = "outputs/contest_signal_model/contest_library.json"
    contest_library = []
    if os.path.exists(library_path):
        try:
            with open(library_path, "r", encoding="utf-8") as f:
                contest_library = json.load(f)
        except:
            pass
            
    # List files in data/
    metadata = file_manager.load_file_metadata()
    data_files = [f for f in metadata.keys()]
    
    if not data_files:
        st.info("No files found in data folder. Upload files in the Central File Manager first.")
    else:
        st.markdown("#### Add Contest to Library")
        source_file = st.selectbox("Select Contest Source File", [os.path.join("data", f) for f in data_files], key="lib_source_file")
        
        selected_sheet = None
        all_cols = []
        if source_file:
            if source_file.lower().endswith((".xlsx", ".xls")):
                try:
                    import pandas as pd
                    xl = pd.ExcelFile(source_file)
                    sheet_names = xl.sheet_names
                    
                    sheet_labels = []
                    sheet_to_label = {}
                    label_to_sheet = {}
                    
                    import openpyxl
                    wb = openpyxl.load_workbook(source_file, read_only=True, data_only=True)
                    for sh_name in sheet_names:
                        label = sh_name
                        if sh_name in wb.sheetnames:
                            sh = wb[sh_name]
                            first_val = None
                            for row in sh.iter_rows(max_row=1, max_col=1, values_only=True):
                                if row:
                                    first_val = row[0]
                            if first_val:
                                label = f"{sh_name} - {str(first_val).strip()}"
                        sheet_labels.append(label)
                        sheet_to_label[sh_name] = label
                        label_to_sheet[label] = sh_name
                    
                    selected_label = st.selectbox("Select Sheet/Table", sheet_labels, key="lib_selected_sheet_label")
                    selected_sheet = label_to_sheet.get(selected_label)
                except Exception as e:
                    print(f"Error reading Excel sheet names: {e}")
                    selected_sheet = None
            
            try:
                import contest_manager
                res_load = contest_manager.inspect_and_load_file(source_file, sheet_name=selected_sheet)
                if res_load["status"] == "success":
                    all_cols = list(res_load["df"].columns)
            except Exception as e:
                print(f"Error loading file columns: {e}")
        
        c_id = st.text_input("Unique Contest ID", value="contest_" + str(len(contest_library) + 1))
        c_name = st.text_input("Contest Name (display name)", value="")
        c_date = st.text_input("Contest Election Date (YYYY-MM-DD)", value="")
        c_year = st.text_input("Contest Election Year", value="")
        
        c_type = st.selectbox("Contest Classification Type", contest_type_opts, key="lib_contest_type")
        c_scope = st.selectbox("Contest Scope Type", scope_type_opts, key="lib_scope_type")
        c_scope_field = st.text_input("Contest Scope Field (e.g. sd or city)", value="")
        c_scope_value = st.text_input("Contest Scope Value (e.g. 4 or SANTA ROSA)", value="")
        
        def_prec_idx = 0
        prec_opts = ["Precinct"] + all_cols if all_cols else ["Precinct"]
        if all_cols:
            for p_c in ["Precinct", "PREC", "PrecinctName", "PREC_JOIN"]:
                if p_c in all_cols:
                    def_prec_idx = prec_opts.index(p_c)
                    break
        c_prec_col = st.selectbox("Precinct Identifier Column", prec_opts, index=def_prec_idx)
        
        c_cross = st.checkbox("Uses Official Sonoma Crosswalk", value=False)
        c_reg_pdf_val = ""
        c_voting_pdf_val = ""
        if c_cross:
            pdf_files = [f for f in data_files if f.lower().endswith(".pdf")]
            c_reg_pdf = st.selectbox("Select Regular-to-Voting Crosswalk PDF (e.g. ewmr010)", ["None"] + pdf_files)
            c_voting_pdf = st.selectbox("Select Voting-to-Regular Crosswalk PDF (e.g. ewmr008)", ["None"] + pdf_files)
            c_reg_pdf_val = "" if c_reg_pdf == "None" else os.path.join("data", c_reg_pdf)
            c_voting_pdf_val = "" if c_voting_pdf == "None" else os.path.join("data", c_voting_pdf)
            
        c_weight = st.slider("Contest Scoring Weight", 0.0, 5.0, 1.0, 0.1)
        c_cross_weight_label = st.slider("Contest Confidence Weight", 0.0, 1.0, 1.0, 0.05)
        c_notes = st.text_input("Contest Notes", value="")
        
        def calculate_file_hash(file_path):
            if not os.path.exists(file_path):
                return ""
            hash_sha256 = hashlib.sha256()
            try:
                with open(file_path, "rb") as f:
                    for chunk in iter(lambda: f.read(4096), b""):
                        hash_sha256.update(chunk)
                return hash_sha256.hexdigest()
            except Exception:
                return ""

        if st.button("➕ Add Contest to Library"):
            if not c_name:
                st.error("Contest Name is required.")
            else:
                import hashlib
                h = calculate_file_hash(source_file)
                new_entry = {
                    "contest_id": c_id,
                    "contest_name": c_name,
                    "election_date": c_date,
                    "election_year": c_year,
                    "contest_type": c_type,
                    "source_file": source_file,
                    "source_file_hash": h,
                    "sheet_name": selected_sheet,
                    "precinct_column": c_prec_col,
                    "scope_type": c_scope,
                    "scope_field": c_scope_field,
                    "scope_value": c_scope_value,
                    "scope_status": "valid",
                    "uses_official_crosswalk": c_cross,
                    "crosswalk_source": "canonical_crosswalk" if c_cross else "",
                    "crosswalk_reg_to_voting_file": c_reg_pdf_val,
                    "crosswalk_voting_to_reg_file": c_voting_pdf_val,
                    "coverage_rate": 100.0,
                    "enabled": True,
                    "contest_weight": c_weight,
                    "confidence_weight": c_cross_weight_label,
                    "notes": c_notes
                }
                contest_library.append(new_entry)
                os.makedirs("outputs/contest_signal_model", exist_ok=True)
                with open(library_path, "w", encoding="utf-8") as f:
                    json.dump(contest_library, f, indent=2)
                st.success(f"Contest '{c_name}' added to library!")
                st.rerun()

        # Render library table and allow delete
        if contest_library:
            st.markdown("#### Library Contests")
            lib_df = pd.DataFrame(contest_library)
            st.dataframe(lib_df[["contest_id", "contest_name", "contest_type", "enabled", "contest_weight"]], use_container_width=True)
            
            # Select and Delete
            to_delete = st.selectbox("Select Contest to Delete", [c["contest_id"] for c in contest_library])
            if st.button("🗑️ Delete Selected Contest"):
                contest_library = [c for c in contest_library if c["contest_id"] != to_delete]
                with open(library_path, "w", encoding="utf-8") as f:
                    json.dump(contest_library, f, indent=2)
                st.success(f"Deleted contest {to_delete}!")
                st.rerun()

    st.markdown("---")
    
    # 3. Column Classification Table Panel
    st.subheader("📊 Column Classification Matrix")
    matrix_path = "outputs/contest_signal_model/contest_column_classification_matrix.csv"
    matrix_df = pd.DataFrame()
    if os.path.exists(matrix_path):
        try:
            matrix_df = pd.read_csv(matrix_path)
        except:
            pass
            
    if not contest_library:
        st.info("Add a contest to the library first to classify columns.")
    else:
        selected_lib_id = st.selectbox("Select Contest to Classify", [c["contest_id"] for c in contest_library])
        selected_contest = [c for c in contest_library if c["contest_id"] == selected_lib_id][0]
        
        # Load contest columns
        c_src = selected_contest["source_file"]
        sh_name = selected_contest.get("sheet_name", None)
        if os.path.exists(c_src):
            try:
                import contest_manager
                res_load = contest_manager.inspect_and_load_file(c_src, sheet_name=sh_name)
                if res_load["status"] == "success":
                    all_cols = list(res_load["df"].columns)
                else:
                    all_cols = []
                    st.error(f"Failed to read file: {res_load['message']}")
            except Exception as e:
                all_cols = []
                st.error(f"Failed to read file columns: {e}")
        else:
            all_cols = []
            
        if all_cols:
            col_to_edit = st.selectbox("Select Column to Classify", all_cols)
            
            # Default values from existing matrix
            existing_row = {}
            if not matrix_df.empty:
                matches = matrix_df[(matrix_df["contest_id"] == selected_lib_id) & (matrix_df["original_column_name"] == col_to_edit)]
                if not matches.empty:
                    existing_row = matches.iloc[0].to_dict()
                    
            sig_types = ["support", "opposition", "turnout", "persuasion", "issue_alignment", "partisan_baseline", "total_votes", "registered_voters", "denominator", "ignore", "unknown"]
            def_sig_idx = sig_types.index(existing_row.get("mapped_signal_type")) if existing_row.get("mapped_signal_type") in sig_types else 10
            mapped_signal_type = st.selectbox("Signal Type", sig_types, index=def_sig_idx)
            
            mapped_side = st.text_input("Mapped Side Name", value=existing_row.get("mapped_side", ""))
            
            relationships = ["supports_current_campaign", "opposes_current_campaign", "neutral_reference", "turnout_only", "issue_similarity", "ignore", "unknown"]
            def_rel_idx = relationships.index(existing_row.get("current_campaign_relationship")) if existing_row.get("current_campaign_relationship") in relationships else 6
            current_campaign_relationship = st.selectbox("Campaign Relationship", relationships, index=def_rel_idx)
            
            directions = ["higher_supports_current_campaign", "higher_opposes_current_campaign", "higher_indicates_turnout", "higher_indicates_persuasion_opportunity", "reference_only", "ignore"]
            def_dir_idx = directions.index(existing_row.get("directionality")) if existing_row.get("directionality") in directions else 4
            directionality = st.selectbox("Directionality", directions, index=def_dir_idx)
            
            denom_column = st.selectbox("Denominator Column", ["None"] + all_cols, index=(all_cols.index(existing_row.get("denominator_column")) + 1 if existing_row.get("denominator_column") in all_cols else 0))
            denom_type = st.text_input("Denominator Type Label", value=existing_row.get("denominator_type", ""))
            
            include_in_scoring = st.checkbox("Include in Scoring", value=existing_row.get("include_in_scoring", True))
            signal_weight = st.number_input("Signal Weight", value=float(existing_row.get("signal_weight", 1.0)))
            confidence_weight = st.number_input("Confidence Weight", value=float(existing_row.get("confidence_weight", 1.0)))
            notes = st.text_input("Classification Notes", value=existing_row.get("notes", ""))
            
            if st.button("💾 Apply to Column"):
                # Remove old row if exists
                if not matrix_df.empty:
                    matrix_df = matrix_df[~((matrix_df["contest_id"] == selected_lib_id) & (matrix_df["original_column_name"] == col_to_edit))]
                
                new_row = {
                    "contest_id": selected_lib_id,
                    "contest_name": selected_contest["contest_name"],
                    "source_file": selected_contest["source_file"],
                    "original_column_name": col_to_edit,
                    "normalized_column_name": col_to_edit,
                    "user_classification": mapped_signal_type,
                    "mapped_signal_type": mapped_signal_type,
                    "mapped_side": mapped_side,
                    "current_campaign_relationship": current_campaign_relationship,
                    "directionality": directionality,
                    "include_in_scoring": include_in_scoring,
                    "signal_weight": signal_weight,
                    "confidence_weight": confidence_weight,
                    "denominator_column": "" if denom_column == "None" else denom_column,
                    "denominator_type": denom_type,
                    "notes": notes
                }
                
                matrix_df = pd.concat([matrix_df, pd.DataFrame([new_row])], ignore_index=True)
                os.makedirs("outputs/contest_signal_model", exist_ok=True)
                matrix_df.to_csv(matrix_path, index=False)
                st.success(f"Column '{col_to_edit}' classification updated!")
                st.rerun()

            # Render classified columns table
            if not matrix_df.empty:
                c_matrix = matrix_df[matrix_df["contest_id"] == selected_lib_id]
                if not c_matrix.empty:
                    st.markdown("#### Classified Columns for Selected Contest")
                    st.dataframe(c_matrix[["original_column_name", "mapped_signal_type", "current_campaign_relationship", "directionality", "denominator_column", "include_in_scoring"]], use_container_width=True)

    st.markdown("---")
    
    # 4. Preview Validation & Scoring Panel
    st.subheader("🚀 Validation & Preview Scoring")
    if not os.path.exists(profile_path) or not os.path.exists(library_path) or not os.path.exists(matrix_path):
        st.info("Configure campaign profile, library, and classifications to execute validation.")
    else:
        if st.button("⚡ Run Contest Signal Preview"):
            try:
                import contest_signal_model
                
                # Load profile, library, classification
                with open(profile_path, "r", encoding="utf-8") as f:
                    campaign_profile = json.load(f)
                with open(library_path, "r", encoding="utf-8") as f:
                    contest_library = json.load(f)
                matrix_df = pd.read_csv(matrix_path)
                
                # Load baseline production priority precincts
                prod_path = "outputs/final_rankings/production_priority_precincts.csv"
                if not os.path.exists(prod_path):
                    # Fall back to run pipeline without contest file to generate base structure
                    st.info("Production priorities not found. Running baseline pipeline...")
                    result = run_pipeline(
                        weights={'turnout_gap': 0.4, 'competitive_index': 0.4, 'density': 0.2},
                        target_params={'ad': None, 'sd': 4, 'city': None},
                        run_mode="PRODUCTION_MODE",
                        trigger_source="streamlit_ui"
                    )
                
                production_df = pd.read_csv(prod_path)
                
                # Execute math engine
                prec_signals_df = contest_signal_model.calculate_precinct_contest_signals(
                    production_df, contest_library, matrix_df, campaign_profile
                )
                
                agg_df = contest_signal_model.aggregate_precinct_signal_scores(
                    prec_signals_df, contest_library
                )
                
                preview_df = contest_signal_model.generate_preview_rankings(
                    production_df, agg_df, campaign_profile
                )
                
                corr_df = contest_signal_model.generate_correlation_matrix(prec_signals_df)
                
                # Write outputs
                os.makedirs("outputs/contest_signal_model", exist_ok=True)
                prec_signals_df.to_csv("outputs/contest_signal_model/precinct_contest_signal_matrix.csv", index=False)
                agg_df.to_csv("outputs/contest_signal_model/aggregate_precinct_signal_scores.csv", index=False)
                preview_df.to_csv("outputs/contest_signal_model/preview_multi_contest_priority_scores.csv", index=False)
                corr_df.to_csv("outputs/contest_signal_model/contest_signal_correlation_matrix.csv", index=False)
                
                validation_md = contest_signal_model.generate_contest_signal_validation_report(
                    contest_library, matrix_df, prec_signals_df, campaign_profile
                )
                with open("outputs/contest_signal_model/contest_signal_validation_report.md", "w", encoding="utf-8") as f:
                    f.write(validation_md)
                    
                st.success("⚡ Preview Calculation Complete!")
                st.session_state["contest_signal_preview_run"] = True
                st.session_state["latest_output_manifest"] = build_output_manifest(st.session_state.get("latest_run_context"))
                st.rerun()
            except Exception as e:
                st.error(f"Preview run failed: {e}")
                import traceback
                st.code(traceback.format_exc())

        # Render preview calculations if exists
        preview_scores_path = "outputs/contest_signal_model/preview_multi_contest_priority_scores.csv"
        if os.path.exists(preview_scores_path):
            st.markdown("### 🏆 Preview Rankings (Top 25)")
            prev_df = pd.read_csv(preview_scores_path)
            st.dataframe(prev_df.sort_values("Preview_Rank").head(25), use_container_width=True)
            
            # Scatter plot of Support vs Turnout
            st.markdown("### 📊 Targeting Distribution Analyses")
            st.scatter_chart(
                prev_df,
                x="Preview_MultiContest_Support_Score",
                y="Preview_MultiContest_Turnout_Score",
                color="Strategic_Bucket",
                size="Preview_MultiContest_Composite_Score"
            )
            
            # Correlation Matrix Render
            corr_path = "outputs/contest_signal_model/contest_signal_correlation_matrix.csv"
            if os.path.exists(corr_path):
                st.markdown("### 🧬 Historical Contest Correlation Matrix")
                corr_df = pd.read_csv(corr_path)
                st.dataframe(corr_df, use_container_width=True)
                
            # Validation summary report Markdown
            val_report_path = "outputs/contest_signal_model/contest_signal_validation_report.md"
            if os.path.exists(val_report_path):
                st.markdown("### 📋 Contest Signal Validation Diagnostic Summary")
                with open(val_report_path, "r", encoding="utf-8") as f:
                    st.markdown(f.read())



